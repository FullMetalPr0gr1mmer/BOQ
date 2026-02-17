"""
DU RPA Logistics Route - Full CRUD operations for DU RPA Logistics

Features:
- Project CRUD with unique PO#
- Description CRUD with calculated stats
- Invoice upload via CSV with validation
- Statistics and analytics endpoints
"""

import json
import logging
import zipfile
import pandas as pd
import openpyxl
from io import StringIO, BytesIO
from datetime import datetime, timedelta
from pathlib import Path
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query, status, Request, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import or_, func
from typing import Optional, List

# Configure logging
logger = logging.getLogger(__name__)

from APIs.Core import get_db, get_current_user
from Models.DU.DU_RPA_Logistics import (
    DURPAProject,
    DURPADescription,
    DURPAInvoice,
    DURPAInvoiceItem
)
from Models.Admin.User import User
from Models.Admin.AuditLog import AuditLog
from Schemas.DU.DU_RPA_Logistics_Schema import (
    CreateDURPAProject,
    UpdateDURPAProject,
    DURPAProjectOut,
    DURPAProjectWithStats,
    DURPAProjectPagination,
    CreateDURPADescription,
    UpdateDURPADescription,
    DURPADescriptionOut,
    DURPADescriptionWithStats,
    DURPADescriptionPagination,
    CreateDURPAInvoice,
    DURPAInvoiceOut,
    DURPAInvoiceItemOut,
    DURPAInvoicePagination,
    UploadResponse,
    BulkDescriptionUpload
)

duRPALogisticsRoute = APIRouter(tags=["DU RPA Logistics"])


# ===========================
# HELPER FUNCTIONS
# ===========================

async def create_audit_log(
        db: Session,
        user_id: int,
        action: str,
        resource_type: str,
        resource_id: Optional[str] = None,
        resource_name: Optional[str] = None,
        details: Optional[str] = None,
        ip_address: Optional[str] = None,
        user_agent: Optional[str] = None
):
    """Create an audit log entry."""
    audit_log = AuditLog(
        user_id=user_id,
        action=action,
        resource_type=resource_type,
        resource_id=resource_id,
        resource_name=resource_name,
        details=details,
        ip_address=ip_address,
        user_agent=user_agent
    )
    db.add(audit_log)
    db.commit()
    return audit_log


def get_client_ip(request: Request) -> str:
    """Extract client IP from request."""
    forwarded = request.headers.get("X-Forwarded-For")
    if forwarded:
        return forwarded.split(",")[0].strip()
    return request.client.host if request.client else "unknown"


def calculate_description_stats(db: Session, description: DURPADescription) -> DURPADescriptionWithStats:
    """Calculate stats for a description."""
    # Get sum of quantities from invoice items
    actual_qty_billed = db.query(func.coalesce(func.sum(DURPAInvoiceItem.quantity), 0)).filter(
        DURPAInvoiceItem.description_id == description.id
    ).scalar() or 0

    # Calculate values
    po_qty_per_unit = description.po_qty_per_unit or 0
    price_per_unit = description.price_per_unit or 0

    total_po_value = po_qty_per_unit * price_per_unit
    actual_value_billed = actual_qty_billed * price_per_unit
    balance = po_qty_per_unit - actual_qty_billed

    return DURPADescriptionWithStats(
        id=description.id,
        project_id=description.project_id,
        description=description.description,
        po_line_item=description.po_line_item,
        po_qty_as_per_po=description.po_qty_as_per_po,
        po_qty_per_unit=description.po_qty_per_unit,
        price_per_unit=description.price_per_unit,
        total_po_value=total_po_value,
        actual_qty_billed=actual_qty_billed,
        actual_value_billed=actual_value_billed,
        balance=balance,
        created_at=description.created_at,
        updated_at=description.updated_at
    )


def calculate_project_stats(db: Session, project: DURPAProject) -> DURPAProjectWithStats:
    """
    OPTIMIZED: Calculate stats for a project using a single SQL query with aggregations.
    Replaces N+1 query pattern with efficient JOIN and GROUP BY.
    """
    from sqlalchemy import case

    # Single query with LEFT JOIN to aggregate all stats at once
    stats = db.query(
        func.count(func.distinct(DURPADescription.id)).label('description_count'),
        func.count(func.distinct(DURPAInvoice.id)).label('invoice_count'),
        func.coalesce(
            func.sum(
                case(
                    (DURPADescription.po_qty_per_unit.isnot(None),
                     DURPADescription.po_qty_per_unit * func.coalesce(DURPADescription.price_per_unit, 0)),
                    else_=0
                )
            ), 0
        ).label('total_po_value'),
        func.coalesce(
            func.sum(
                DURPAInvoiceItem.quantity * func.coalesce(DURPADescription.price_per_unit, 0)
            ), 0
        ).label('total_billed_value')
    ).outerjoin(
        DURPADescription, DURPADescription.project_id == project.id
    ).outerjoin(
        DURPAInvoice, DURPAInvoice.project_id == project.id
    ).outerjoin(
        DURPAInvoiceItem, DURPAInvoiceItem.description_id == DURPADescription.id
    ).filter(
        or_(
            DURPADescription.project_id == project.id,
            DURPAInvoice.project_id == project.id
        )
    ).first()

    return DURPAProjectWithStats(
        id=project.id,
        po_number=project.po_number,
        category=project.category or 'ppo_based',
        created_at=project.created_at,
        updated_at=project.updated_at,
        description_count=stats.description_count or 0,
        invoice_count=stats.invoice_count or 0,
        total_po_value=float(stats.total_po_value or 0),
        total_billed_value=float(stats.total_billed_value or 0)
    )


def calculate_descriptions_stats_bulk(db: Session, descriptions: List[DURPADescription]) -> List[DURPADescriptionWithStats]:
    """
    OPTIMIZED: Calculate stats for multiple descriptions in a single query.
    Uses GROUP BY to aggregate all quantities at once instead of N queries.
    """
    if not descriptions:
        return []

    # Get all description IDs
    desc_ids = [d.id for d in descriptions]

    # Single query to get aggregated quantities for all descriptions
    qty_aggregates = db.query(
        DURPAInvoiceItem.description_id,
        func.coalesce(func.sum(DURPAInvoiceItem.quantity), 0).label('total_qty')
    ).filter(
        DURPAInvoiceItem.description_id.in_(desc_ids)
    ).group_by(
        DURPAInvoiceItem.description_id
    ).all()

    # Create lookup dict for O(1) access
    qty_lookup = {desc_id: float(qty) for desc_id, qty in qty_aggregates}

    # Build result list using in-memory calculations
    results = []
    for desc in descriptions:
        actual_qty_billed = qty_lookup.get(desc.id, 0)
        po_qty_per_unit = desc.po_qty_per_unit or 0
        price_per_unit = desc.price_per_unit or 0

        total_po_value = po_qty_per_unit * price_per_unit
        actual_value_billed = actual_qty_billed * price_per_unit
        balance = po_qty_per_unit - actual_qty_billed

        results.append(DURPADescriptionWithStats(
            id=desc.id,
            project_id=desc.project_id,
            description=desc.description,
            po_line_item=desc.po_line_item,
            po_qty_as_per_po=desc.po_qty_as_per_po,
            po_qty_per_unit=desc.po_qty_per_unit,
            price_per_unit=desc.price_per_unit,
            total_po_value=total_po_value,
            actual_qty_billed=actual_qty_billed,
            actual_value_billed=actual_value_billed,
            balance=balance,
            created_at=desc.created_at,
            updated_at=desc.updated_at
        ))

    return results


def calculate_projects_stats_bulk(db: Session, projects: List[DURPAProject]) -> List[DURPAProjectWithStats]:
    """
    OPTIMIZED: Calculate stats for multiple projects in a single query.
    Uses GROUP BY to aggregate all project stats at once instead of N queries.
    """
    if not projects:
        return []

    from sqlalchemy import case

    # Get all project IDs
    project_ids = [p.id for p in projects]

    # Single mega-query with GROUP BY to calculate all project stats
    stats_query = db.query(
        DURPADescription.project_id,
        func.count(func.distinct(DURPADescription.id)).label('description_count'),
        func.coalesce(
            func.sum(
                case(
                    (DURPADescription.po_qty_per_unit.isnot(None),
                     DURPADescription.po_qty_per_unit * func.coalesce(DURPADescription.price_per_unit, 0)),
                    else_=0
                )
            ), 0
        ).label('total_po_value'),
        func.coalesce(
            func.sum(
                DURPAInvoiceItem.quantity * func.coalesce(DURPADescription.price_per_unit, 0)
            ), 0
        ).label('total_billed_value')
    ).outerjoin(
        DURPAInvoiceItem, DURPAInvoiceItem.description_id == DURPADescription.id
    ).filter(
        DURPADescription.project_id.in_(project_ids)
    ).group_by(
        DURPADescription.project_id
    ).all()

    # Get invoice counts separately (simpler query)
    invoice_counts = db.query(
        DURPAInvoice.project_id,
        func.count(DURPAInvoice.id).label('invoice_count')
    ).filter(
        DURPAInvoice.project_id.in_(project_ids)
    ).group_by(
        DURPAInvoice.project_id
    ).all()

    # Create lookup dicts for O(1) access
    stats_lookup = {row.project_id: row for row in stats_query}
    invoice_lookup = {row.project_id: row.invoice_count for row in invoice_counts}

    # Build results using lookups
    results = []
    for project in projects:
        stats = stats_lookup.get(project.id)

        results.append(DURPAProjectWithStats(
            id=project.id,
            po_number=project.po_number,
            category=project.category or 'ppo_based',
            created_at=project.created_at,
            updated_at=project.updated_at,
            description_count=stats.description_count if stats else 0,
            invoice_count=invoice_lookup.get(project.id, 0),
            total_po_value=float(stats.total_po_value) if stats else 0,
            total_billed_value=float(stats.total_billed_value) if stats else 0
        ))

    return results


# ===========================
# PROJECT ENDPOINTS
# ===========================

@duRPALogisticsRoute.post("/du-rpa/projects", response_model=DURPAProjectOut)
async def create_project(
        project_data: CreateDURPAProject,
        request: Request,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    """Create a new DU RPA Project with unique PO#."""
    # Check if PO# already exists
    existing = db.query(DURPAProject).filter(DURPAProject.po_number == project_data.po_number).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Project with PO# '{project_data.po_number}' already exists"
        )

    try:
        new_project = DURPAProject(po_number=project_data.po_number, category=project_data.category)
        db.add(new_project)
        db.commit()
        db.refresh(new_project)

        await create_audit_log(
            db=db,
            user_id=current_user.id,
            action="create_du_rpa_project",
            resource_type="du_rpa_project",
            resource_id=str(new_project.id),
            resource_name=new_project.po_number,
            ip_address=get_client_ip(request),
            user_agent=request.headers.get("User-Agent")
        )

        return new_project

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error creating project: {str(e)}"
        )


@duRPALogisticsRoute.get("/du-rpa/projects", response_model=DURPAProjectPagination)
def get_projects(
        skip: int = Query(0, ge=0),
        limit: int = Query(50, ge=1, le=500),
        search: Optional[str] = None,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    """Get all DU RPA Projects with pagination. OPTIMIZED with bulk stats calculation."""
    try:
        query = db.query(DURPAProject)

        if search:
            query = query.filter(DURPAProject.po_number.ilike(f"%{search}%"))

        total = query.count()
        projects = query.order_by(DURPAProject.created_at.desc()).offset(skip).limit(limit).all()

        # OPTIMIZED: Calculate stats for all projects in bulk (single query vs N queries)
        records = calculate_projects_stats_bulk(db, projects)

        return DURPAProjectPagination(records=records, total=total)

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error retrieving projects: {str(e)}"
        )


@duRPALogisticsRoute.get("/du-rpa/projects/{project_id}", response_model=DURPAProjectWithStats)
def get_project(
        project_id: int,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    """Get a specific DU RPA Project by ID."""
    project = db.query(DURPAProject).filter(DURPAProject.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    return calculate_project_stats(db, project)


@duRPALogisticsRoute.put("/du-rpa/projects/{project_id}", response_model=DURPAProjectOut)
async def update_project(
        project_id: int,
        project_data: UpdateDURPAProject,
        request: Request,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    """Update an existing DU RPA Project."""
    project = db.query(DURPAProject).filter(DURPAProject.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    if project_data.po_number:
        # Check if new PO# already exists
        existing = db.query(DURPAProject).filter(
            DURPAProject.po_number == project_data.po_number,
            DURPAProject.id != project_id
        ).first()
        if existing:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Project with PO# '{project_data.po_number}' already exists"
            )
        project.po_number = project_data.po_number
    if project_data.category is not None:
        project.category = project_data.category

    try:
        db.commit()
        db.refresh(project)

        await create_audit_log(
            db=db,
            user_id=current_user.id,
            action="update_du_rpa_project",
            resource_type="du_rpa_project",
            resource_id=str(project.id),
            resource_name=project.po_number,
            ip_address=get_client_ip(request),
            user_agent=request.headers.get("User-Agent")
        )

        return project

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error updating project: {str(e)}"
        )


@duRPALogisticsRoute.delete("/du-rpa/projects/{project_id}")
async def delete_project(
        project_id: int,
        request: Request,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    """Delete a DU RPA Project and all related data."""
    project = db.query(DURPAProject).filter(DURPAProject.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    try:
        po_number = project.po_number
        db.delete(project)
        db.commit()

        await create_audit_log(
            db=db,
            user_id=current_user.id,
            action="delete_du_rpa_project",
            resource_type="du_rpa_project",
            resource_id=str(project_id),
            resource_name=po_number,
            ip_address=get_client_ip(request),
            user_agent=request.headers.get("User-Agent")
        )

        return {"message": f"Project '{po_number}' deleted successfully"}

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error deleting project: {str(e)}"
        )


# ===========================
# DESCRIPTION ENDPOINTS
# ===========================

@duRPALogisticsRoute.post("/du-rpa/projects/{project_id}/descriptions", response_model=DURPADescriptionWithStats)
async def create_description(
        project_id: int,
        desc_data: CreateDURPADescription,
        request: Request,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    """Create a new description for a project."""
    project = db.query(DURPAProject).filter(DURPAProject.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    try:
        new_desc = DURPADescription(
            project_id=project_id,
            description=desc_data.description,
            po_line_item=desc_data.po_line_item,
            po_qty_as_per_po=desc_data.po_qty_as_per_po,
            po_qty_per_unit=desc_data.po_qty_per_unit,
            price_per_unit=desc_data.price_per_unit
        )
        db.add(new_desc)
        db.commit()
        db.refresh(new_desc)

        await create_audit_log(
            db=db,
            user_id=current_user.id,
            action="create_du_rpa_description",
            resource_type="du_rpa_description",
            resource_id=str(new_desc.id),
            resource_name=new_desc.description[:100],
            details=json.dumps({"project_id": project_id}),
            ip_address=get_client_ip(request),
            user_agent=request.headers.get("User-Agent")
        )

        return calculate_description_stats(db, new_desc)

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error creating description: {str(e)}"
        )


@duRPALogisticsRoute.post("/du-rpa/projects/{project_id}/descriptions/bulk", response_model=UploadResponse)
async def bulk_create_descriptions(
        project_id: int,
        bulk_data: BulkDescriptionUpload,
        request: Request,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    """Bulk create descriptions for a project."""
    project = db.query(DURPAProject).filter(DURPAProject.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    inserted = 0
    errors = []

    for idx, desc_data in enumerate(bulk_data.descriptions):
        try:
            new_desc = DURPADescription(
                project_id=project_id,
                description=desc_data.description,
                po_line_item=desc_data.po_line_item,
                po_qty_as_per_po=desc_data.po_qty_as_per_po,
                po_qty_per_unit=desc_data.po_qty_per_unit,
                price_per_unit=desc_data.price_per_unit
            )
            db.add(new_desc)
            inserted += 1
        except Exception as e:
            errors.append(f"Row {idx + 1}: {str(e)}")

    try:
        db.commit()

        await create_audit_log(
            db=db,
            user_id=current_user.id,
            action="bulk_create_du_rpa_descriptions",
            resource_type="du_rpa_description",
            resource_id=str(project_id),
            resource_name=project.po_number,
            details=json.dumps({"inserted": inserted, "errors": len(errors)}),
            ip_address=get_client_ip(request),
            user_agent=request.headers.get("User-Agent")
        )

        return UploadResponse(
            inserted=inserted,
            errors=errors,
            message=f"Successfully inserted {inserted} descriptions"
        )

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error bulk creating descriptions: {str(e)}"
        )


@duRPALogisticsRoute.post("/du-rpa/projects/{project_id}/descriptions/upload-csv", response_model=UploadResponse)
async def upload_descriptions_csv(
        project_id: int,
        file: UploadFile = File(...),
        request: Request = None,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    """
    Upload descriptions via CSV.

    Supports two formats:

    1) Row-based (standard):
       Columns: description, po_line_item, po_qty_as_per_po, po_qty_per_unit, price_per_unit

    2) Transposed (PO tracker):
       Row 1 col F+: PO Line Item values
       Row 2 col F+: PO Qty (as per PO)
       Row 3 col F+: PO Qty (per unit)
       Row 4 col F+: Price per Unit
       Row 11 col F+: Description names
       Auto-detected when cell A1 starts with "PO#".
    """
    project = db.query(DURPAProject).filter(DURPAProject.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    if not file.filename.endswith('.csv'):
        raise HTTPException(status_code=400, detail="File must be a CSV")

    try:
        content = await file.read()
        # Try multiple encodings
        try:
            csv_content = content.decode('utf-8-sig')
        except UnicodeDecodeError:
            try:
                csv_content = content.decode('windows-1252')
            except UnicodeDecodeError:
                csv_content = content.decode('iso-8859-1')

        inserted = 0
        errors = []
        bulk_data = []

        # Read without header to detect format
        raw_df = pd.read_csv(StringIO(csv_content), header=None)
        is_transposed = str(raw_df.iloc[0, 0]).strip().upper().startswith('PO#')

        if is_transposed:
            # Transposed PO tracker format
            # Row 0: PO Line Item (col F+)
            # Row 1: PO Qty (as per PO)
            # Row 2: PO Qty (per unit)
            # Row 3: Price per Unit
            # Row 10: Description names
            if len(raw_df) < 11:
                raise HTTPException(status_code=400, detail="Transposed CSV must have at least 11 rows")

            # Helper: safely parse a float cell, returning None for NaN/empty
            def parse_float(val):
                if pd.isna(val):
                    return None
                try:
                    return float(val)
                except (ValueError, TypeError):
                    return None

            for col_idx in range(5, len(raw_df.columns)):
                try:
                    desc_text = str(raw_df.iloc[10, col_idx]).strip()
                    if not desc_text or desc_text == 'nan':
                        continue

                    # PO Line Item — skip columns with no line item (trailing metadata columns)
                    li_val = raw_df.iloc[0, col_idx]
                    if pd.isna(li_val):
                        continue
                    po_line_item = str(li_val).strip() or None

                    # PO Qty (as per PO)
                    po_qty_as_per_po = parse_float(raw_df.iloc[1, col_idx])

                    # PO Qty (per unit)
                    po_qty_per_unit = parse_float(raw_df.iloc[2, col_idx])

                    # Price per Unit (strip currency formatting: $, commas, spaces)
                    price_per_unit = None
                    price_raw = raw_df.iloc[3, col_idx]
                    if pd.notna(price_raw):
                        try:
                            price_str = str(price_raw).replace('$', '').replace(',', '').strip()
                            price_per_unit = float(price_str) if price_str else None
                        except (ValueError, TypeError):
                            pass

                    bulk_data.append({
                        'project_id': project_id,
                        'description': desc_text,
                        'po_line_item': po_line_item,
                        'po_qty_as_per_po': po_qty_as_per_po,
                        'po_qty_per_unit': po_qty_per_unit,
                        'price_per_unit': price_per_unit
                    })
                    inserted += 1

                except Exception as e:
                    errors.append(f"Column {col_idx + 1}: {str(e)}")
        else:
            # Standard row-based format
            df = pd.read_csv(StringIO(csv_content))
            df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')

            if 'description' not in df.columns:
                raise HTTPException(status_code=400, detail="CSV must have a 'description' column")

            for idx, row in df.iterrows():
                try:
                    desc_text = str(row.get('description', '')).strip()
                    if not desc_text:
                        errors.append(f"Row {idx + 2}: Empty description")
                        continue

                    bulk_data.append({
                        'project_id': project_id,
                        'description': desc_text,
                        'po_line_item': str(row.get('po_line_item', '')).strip() or None,
                        'po_qty_as_per_po': float(row['po_qty_as_per_po']) if pd.notna(row.get('po_qty_as_per_po')) else None,
                        'po_qty_per_unit': float(row['po_qty_per_unit']) if pd.notna(row.get('po_qty_per_unit')) else None,
                        'price_per_unit': float(row['price_per_unit']) if pd.notna(row.get('price_per_unit')) else None
                    })
                    inserted += 1

                except Exception as e:
                    errors.append(f"Row {idx + 2}: {str(e)}")

        # Single bulk insert
        if bulk_data:
            db.bulk_insert_mappings(DURPADescription, bulk_data)

        db.commit()

        await create_audit_log(
            db=db,
            user_id=current_user.id,
            action="upload_du_rpa_descriptions_csv",
            resource_type="du_rpa_description",
            resource_id=str(project_id),
            resource_name=file.filename,
            details=json.dumps({"inserted": inserted, "errors": len(errors), "format": "transposed" if is_transposed else "standard"}),
            ip_address=get_client_ip(request),
            user_agent=request.headers.get("User-Agent")
        )

        return UploadResponse(
            inserted=inserted,
            errors=errors[:10],
            message=f"Successfully inserted {inserted} descriptions. {len(errors)} errors."
        )

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error processing CSV: {str(e)}"
        )


@duRPALogisticsRoute.get("/du-rpa/projects/{project_id}/descriptions", response_model=DURPADescriptionPagination)
def get_descriptions(
        project_id: int,
        skip: int = Query(0, ge=0),
        limit: int = Query(50, ge=1, le=500),
        search: Optional[str] = None,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    """Get all descriptions for a project with calculated stats. OPTIMIZED with bulk calculation."""
    project = db.query(DURPAProject).filter(DURPAProject.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    try:
        query = db.query(DURPADescription).filter(DURPADescription.project_id == project_id)

        if search:
            query = query.filter(
                or_(
                    DURPADescription.description.ilike(f"%{search}%"),
                    DURPADescription.po_line_item.ilike(f"%{search}%")
                )
            )

        total = query.count()
        descriptions = query.order_by(DURPADescription.id).offset(skip).limit(limit).all()

        # OPTIMIZED: Calculate stats for all descriptions in bulk (1 query vs N queries)
        records = calculate_descriptions_stats_bulk(db, descriptions)

        return DURPADescriptionPagination(records=records, total=total)

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error retrieving descriptions: {str(e)}"
        )


@duRPALogisticsRoute.get("/du-rpa/descriptions/{description_id}", response_model=DURPADescriptionWithStats)
def get_description(
        description_id: int,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    """Get a specific description by ID."""
    description = db.query(DURPADescription).filter(DURPADescription.id == description_id).first()
    if not description:
        raise HTTPException(status_code=404, detail="Description not found")

    return calculate_description_stats(db, description)


@duRPALogisticsRoute.put("/du-rpa/descriptions/{description_id}", response_model=DURPADescriptionWithStats)
async def update_description(
        description_id: int,
        desc_data: UpdateDURPADescription,
        request: Request,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    """Update an existing description."""
    description = db.query(DURPADescription).filter(DURPADescription.id == description_id).first()
    if not description:
        raise HTTPException(status_code=404, detail="Description not found")

    try:
        update_data = desc_data.dict(exclude_unset=True)
        for field, value in update_data.items():
            if value is not None:
                setattr(description, field, value)

        db.commit()
        db.refresh(description)

        await create_audit_log(
            db=db,
            user_id=current_user.id,
            action="update_du_rpa_description",
            resource_type="du_rpa_description",
            resource_id=str(description.id),
            resource_name=description.description[:100],
            details=json.dumps(update_data),
            ip_address=get_client_ip(request),
            user_agent=request.headers.get("User-Agent")
        )

        return calculate_description_stats(db, description)

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error updating description: {str(e)}"
        )


@duRPALogisticsRoute.delete("/du-rpa/descriptions/{description_id}")
async def delete_description(
        description_id: int,
        request: Request,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    """Delete a description and its related invoice items."""
    description = db.query(DURPADescription).filter(DURPADescription.id == description_id).first()
    if not description:
        raise HTTPException(status_code=404, detail="Description not found")

    try:
        desc_text = description.description[:100]

        # First delete related invoice items (since we use NO ACTION on FK)
        db.query(DURPAInvoiceItem).filter(DURPAInvoiceItem.description_id == description_id).delete(synchronize_session=False)

        db.delete(description)
        db.commit()

        await create_audit_log(
            db=db,
            user_id=current_user.id,
            action="delete_du_rpa_description",
            resource_type="du_rpa_description",
            resource_id=str(description_id),
            resource_name=desc_text,
            ip_address=get_client_ip(request),
            user_agent=request.headers.get("User-Agent")
        )

        return {"message": "Description deleted successfully"}

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error deleting description: {str(e)}"
        )


# ===========================
# INVOICE ENDPOINTS
# ===========================

@duRPALogisticsRoute.post("/du-rpa/projects/{project_id}/invoices/upload-csv", response_model=UploadResponse)
async def upload_invoices_csv(
        project_id: int,
        file: UploadFile = File(...),
        request: Request = None,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    """
    Upload invoices via CSV (Transposed Format - One row per invoice item).

    Expected columns (case-insensitive):
    1. PO# - To match with project
    2. New PO number
    3. PR #
    4. PPO# - Unique per invoice (prevents duplicates)
    5. Site ID
    6. Model
    7. LI# - Line Item number
    8. Descriptions - Must match existing descriptions
    9. QTY - Quantity
    10. Unit price - Must match description's price_per_unit
    11. PAC date
    12. SAP Invoice
    13. Invoice Date
    14. Customer Invoice

    Each row represents one invoice item. Rows with the same PPO# are grouped into one invoice.
    """
    project = db.query(DURPAProject).filter(DURPAProject.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    if not file.filename.endswith('.csv'):
        raise HTTPException(status_code=400, detail="File must be a CSV")

    try:
        content = await file.read()
        # Try multiple encodings
        try:
            csv_content = content.decode('utf-8-sig')
        except UnicodeDecodeError:
            try:
                csv_content = content.decode('windows-1252')
            except UnicodeDecodeError:
                csv_content = content.decode('iso-8859-1')

        df = pd.read_csv(StringIO(csv_content), header=2)

        # Normalize column names
        df.columns = df.columns.str.strip()

        # Map columns (case-insensitive)
        col_mapping = {}
        for col in df.columns:
            col_lower = col.strip().lower().replace(' ', '_').replace('#', '_').replace('__', '_').strip('_')
            if col_lower in ['po', 'po_number']:
                col_mapping['po_number'] = col
            elif col_lower in ['new_po_number', 'new_po']:
                col_mapping['new_po_number'] = col
            elif col_lower in ['pr', 'pr_number']:
                col_mapping['pr_number'] = col
            elif col_lower in ['ppo', 'ppo_number']:
                col_mapping['ppo_number'] = col
            elif col_lower in ['site_id', 'site']:
                col_mapping['site_id'] = col
            elif col_lower in ['model', 'model_scope', 'model/scope']:
                col_mapping['model'] = col
            elif col_lower in ['li', 'li_number']:
                col_mapping['li_number'] = col
            elif col_lower in ['descriptions', 'description']:
                col_mapping['description'] = col
            elif col_lower in ['qty', 'quantity']:
                col_mapping['qty'] = col
            elif col_lower in ['unit_price', 'price']:
                col_mapping['unit_price'] = col
            elif col_lower in ['pac_date', 'pac']:
                col_mapping['pac_date'] = col
            elif col_lower in ['sap_invoice', 'sap_invoice_number']:
                col_mapping['sap_invoice'] = col
            elif col_lower in ['invoice_date']:
                col_mapping['invoice_date'] = col
            elif col_lower in ['customer_invoice', 'customer_invoice_number', 'customer_invoice_']:
                col_mapping['customer_invoice'] = col
            elif col_lower in ['prf_%', 'prf_', 'prf_percentage']:
                col_mapping['prf_percentage'] = col
            elif col_lower in ['vat_rate']:
                col_mapping['vat_rate'] = col

        # Validate required columns
        required = ['ppo_number', 'description', 'qty']
        missing = [col for col in required if col not in col_mapping]
        if missing:
            raise HTTPException(status_code=400, detail=f"Missing required columns: {', '.join(missing)}")

        # Verify PO# matches project
        if 'po_number' in col_mapping:
            csv_po = str(df[col_mapping['po_number']].iloc[0]).strip()
            if csv_po.lower() != project.po_number.lower():
                raise HTTPException(status_code=400, detail=f"PO# mismatch: CSV has '{csv_po}', expected '{project.po_number}'")

        # Get all descriptions for this project
        descriptions = db.query(DURPADescription).filter(DURPADescription.project_id == project_id).all()
        desc_map = {d.description.strip().lower(): d for d in descriptions}

        # Group rows by PPO#
        invoice_groups = {}
        errors = []
        skipped_invoices = set()

        for idx, row in df.iterrows():
            try:
                ppo_number = str(row[col_mapping['ppo_number']]).strip()

                # Check if this invoice was already skipped due to errors
                if ppo_number in skipped_invoices:
                    continue

                # Parse description and validate
                description_text = str(row[col_mapping['description']]).strip()
                desc_key = description_text.lower()

                if desc_key not in desc_map:
                    errors.append(f"Row {idx + 2}: Unknown description '{description_text}' for PPO# {ppo_number}. Entire invoice declined.")
                    skipped_invoices.add(ppo_number)
                    # Remove this invoice group if it was already started
                    if ppo_number in invoice_groups:
                        del invoice_groups[ppo_number]
                    continue

                desc = desc_map[desc_key]

                # Validate unit price if provided
                unit_price = None
                if 'unit_price' in col_mapping and pd.notna(row.get(col_mapping['unit_price'])):
                    try:
                        unit_price = float(row[col_mapping['unit_price']])
                        if desc.price_per_unit is not None and abs(unit_price - desc.price_per_unit) > 0.01:
                            errors.append(f"Row {idx + 2}: Unit price mismatch for '{description_text}'. Expected {desc.price_per_unit}, got {unit_price}. Entire invoice declined.")
                            skipped_invoices.add(ppo_number)
                            if ppo_number in invoice_groups:
                                del invoice_groups[ppo_number]
                            continue
                    except ValueError:
                        errors.append(f"Row {idx + 2}: Invalid unit price. Entire invoice declined.")
                        skipped_invoices.add(ppo_number)
                        if ppo_number in invoice_groups:
                            del invoice_groups[ppo_number]
                        continue

                # Parse quantity
                try:
                    quantity = float(row[col_mapping['qty']])
                except ValueError:
                    errors.append(f"Row {idx + 2}: Invalid quantity. Entire invoice declined.")
                    skipped_invoices.add(ppo_number)
                    if ppo_number in invoice_groups:
                        del invoice_groups[ppo_number]
                    continue

                # Initialize invoice group if not exists
                if ppo_number not in invoice_groups:
                    # Parse dates
                    invoice_date = None
                    if 'invoice_date' in col_mapping and pd.notna(row.get(col_mapping['invoice_date'])):
                        try:
                            invoice_date = pd.to_datetime(row[col_mapping['invoice_date']]).date()
                        except:
                            pass

                    # Parse PRF %
                    prf_pct = None
                    if 'prf_percentage' in col_mapping and pd.notna(row.get(col_mapping['prf_percentage'])):
                        try:
                            prf_pct = float(str(row[col_mapping['prf_percentage']]).replace('%', '').strip())
                        except (ValueError, TypeError):
                            pass

                    # Parse VAT Rate
                    vat = None
                    if 'vat_rate' in col_mapping and pd.notna(row.get(col_mapping['vat_rate'])):
                        try:
                            vat = float(str(row[col_mapping['vat_rate']]).replace('%', '').strip())
                        except (ValueError, TypeError):
                            pass

                    invoice_groups[ppo_number] = {
                        'new_po_number': str(row.get(col_mapping.get('new_po_number', ''), '')).strip() or None,
                        'pr_number': str(row.get(col_mapping.get('pr_number', ''), '')).strip() or None,
                        'site_id': str(row.get(col_mapping.get('site_id', ''), '')).strip() or None,
                        'model': str(row.get(col_mapping.get('model', ''), '')).strip() or None,
                        'sap_invoice': str(row.get(col_mapping.get('sap_invoice', ''), '')).strip() or None,
                        'invoice_date': invoice_date,
                        'customer_invoice': str(row.get(col_mapping.get('customer_invoice', ''), '')).strip() or None,
                        'prf_percentage': prf_pct,
                        'vat_rate': vat,
                        'items': []
                    }

                # Parse PAC date for item
                pac_date = None
                if 'pac_date' in col_mapping and pd.notna(row.get(col_mapping['pac_date'])):
                    try:
                        pac_date = pd.to_datetime(row[col_mapping['pac_date']]).date()
                    except:
                        pass

                # Add item to invoice group
                invoice_groups[ppo_number]['items'].append({
                    'li_number': str(row.get(col_mapping.get('li_number', ''), '')).strip() or None,
                    'description_id': desc.id,
                    'quantity': quantity,
                    'unit_price': unit_price,
                    'pac_date': pac_date
                })

            except Exception as e:
                errors.append(f"Row {idx + 2}: {str(e)}")

        # OPTIMIZED: Check for existing PPO#s in single query
        ppo_numbers = list(invoice_groups.keys())
        existing_ppos = db.query(DURPAInvoice.ppo_number).filter(
            DURPAInvoice.ppo_number.in_(ppo_numbers)
        ).all()
        existing_ppo_set = {ppo[0] for ppo in existing_ppos}

        # Filter out existing PPO#s
        for ppo in existing_ppo_set:
            errors.append(f"PPO# {ppo}: Already exists. Skipping.")
            del invoice_groups[ppo]

        # OPTIMIZED: Prepare bulk insert data
        invoice_bulk_data = []
        items_by_ppo = {}  # Will hold items keyed by PPO# temporarily
        inserted_invoices = 0
        inserted_items = 0

        for ppo_number, invoice_data in invoice_groups.items():
            invoice_bulk_data.append({
                'project_id': project_id,
                'ppo_number': ppo_number,
                'new_po_number': invoice_data['new_po_number'],
                'pr_number': invoice_data['pr_number'],
                'site_id': invoice_data['site_id'],
                'model': invoice_data['model'],
                'sap_invoice_number': invoice_data['sap_invoice'],
                'invoice_date': invoice_data['invoice_date'],
                'customer_invoice_number': invoice_data['customer_invoice'],
                'prf_percentage': invoice_data['prf_percentage'],
                'vat_rate': invoice_data['vat_rate']
            })
            items_by_ppo[ppo_number] = invoice_data['items']
            inserted_invoices += 1

        # OPTIMIZED: Bulk insert invoices
        if invoice_bulk_data:
            db.bulk_insert_mappings(DURPAInvoice, invoice_bulk_data)
            db.flush()

            # Get the newly created invoice IDs (need to query back by PPO#)
            new_invoices = db.query(DURPAInvoice.id, DURPAInvoice.ppo_number).filter(
                DURPAInvoice.ppo_number.in_(ppo_numbers)
            ).all()
            invoice_id_map = {ppo: inv_id for inv_id, ppo in new_invoices}

            # OPTIMIZED: Prepare bulk insert for invoice items
            items_bulk_data = []
            for ppo_number, items in items_by_ppo.items():
                invoice_id = invoice_id_map.get(ppo_number)
                if not invoice_id:
                    continue

                for item_data in items:
                    items_bulk_data.append({
                        'invoice_id': invoice_id,
                        'description_id': item_data['description_id'],
                        'li_number': item_data['li_number'],
                        'quantity': item_data['quantity'],
                        'unit_price': item_data['unit_price'],
                        'pac_date': item_data['pac_date']
                    })
                    inserted_items += 1

            # OPTIMIZED: Bulk insert invoice items
            if items_bulk_data:
                db.bulk_insert_mappings(DURPAInvoiceItem, items_bulk_data)

        db.commit()

        await create_audit_log(
            db=db,
            user_id=current_user.id,
            action="upload_du_rpa_invoices_csv",
            resource_type="du_rpa_invoice",
            resource_id=str(project_id),
            resource_name=file.filename,
            details=json.dumps({"inserted_invoices": inserted_invoices, "inserted_items": inserted_items, "errors": len(errors)}),
            ip_address=get_client_ip(request),
            user_agent=request.headers.get("User-Agent")
        )

        return UploadResponse(
            inserted=inserted_invoices,
            errors=errors[:50],  # Return first 50 errors
            message=f"Successfully inserted {inserted_invoices} invoices with {inserted_items} items. {len(errors)} errors."
        )

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error processing CSV: {str(e)}"
        )


@duRPALogisticsRoute.get("/du-rpa/projects/{project_id}/invoices", response_model=DURPAInvoicePagination)
def get_invoices(
        project_id: int,
        skip: int = Query(0, ge=0),
        limit: int = Query(50, ge=1, le=500),
        search: Optional[str] = None,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    """Get all invoices for a project. OPTIMIZED: Uses eager loading to prevent N+1 queries."""
    from sqlalchemy.orm import joinedload

    project = db.query(DURPAProject).filter(DURPAProject.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    try:
        # OPTIMIZED: Eager load invoice items and descriptions
        query = db.query(DURPAInvoice).filter(DURPAInvoice.project_id == project_id).options(
            joinedload(DURPAInvoice.items).joinedload(DURPAInvoiceItem.description)
        )

        if search:
            query = query.filter(
                or_(
                    DURPAInvoice.site_id.ilike(f"%{search}%"),
                    DURPAInvoice.sap_invoice_number.ilike(f"%{search}%"),
                    DURPAInvoice.customer_invoice_number.ilike(f"%{search}%")
                )
            )

        total = query.count()

        # OPTIMIZED: All data loaded in ~2 queries instead of 100+
        invoices = query.order_by(DURPAInvoice.created_at.desc()).offset(skip).limit(limit).all()

        # Build response - NO additional queries needed
        records = []
        for inv in invoices:
            items = []
            for item in inv.items:
                items.append(DURPAInvoiceItemOut(
                    id=item.id,
                    invoice_id=item.invoice_id,
                    description_id=item.description_id,
                    li_number=item.li_number,
                    quantity=item.quantity,
                    unit_price=item.unit_price,
                    pac_date=item.pac_date,
                    description_text=item.description.description if item.description else None
                ))

            records.append(DURPAInvoiceOut(
                id=inv.id,
                project_id=inv.project_id,
                ppo_number=inv.ppo_number,
                new_po_number=inv.new_po_number,
                pr_number=inv.pr_number,
                site_id=inv.site_id,
                model=inv.model,
                sap_invoice_number=inv.sap_invoice_number,
                invoice_date=inv.invoice_date,
                customer_invoice_number=inv.customer_invoice_number,
                prf_percentage=inv.prf_percentage,
                vat_rate=inv.vat_rate,
                created_at=inv.created_at,
                items=items
            ))

        return DURPAInvoicePagination(records=records, total=total)

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error retrieving invoices: {str(e)}"
        )


@duRPALogisticsRoute.get("/du-rpa/invoices")
def get_all_invoices(
        skip: int = Query(0, ge=0),
        limit: int = Query(50, ge=1, le=500),
        search: Optional[str] = None,
        po_filter: Optional[str] = None,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    """
    Get all invoices across all projects with pagination and filtering.
    OPTIMIZED: Uses eager loading to prevent N+1 queries.
    """
    from sqlalchemy.orm import joinedload

    try:
        # OPTIMIZED: Eager load relationships to prevent N+1 queries
        query = db.query(DURPAInvoice).join(DURPAProject).options(
            joinedload(DURPAInvoice.project),
            joinedload(DURPAInvoice.items).joinedload(DURPAInvoiceItem.description)
        )

        # Filter by PO# if specified
        if po_filter:
            query = query.filter(DURPAProject.po_number == po_filter)

        # Search across multiple fields (including PPO#)
        if search:
            query = query.filter(
                or_(
                    DURPAInvoice.ppo_number.ilike(f"%{search}%"),
                    DURPAInvoice.site_id.ilike(f"%{search}%"),
                    DURPAInvoice.sap_invoice_number.ilike(f"%{search}%"),
                    DURPAInvoice.customer_invoice_number.ilike(f"%{search}%"),
                    DURPAProject.po_number.ilike(f"%{search}%")
                )
            )

        # Count before applying limit
        total = query.count()

        # OPTIMIZED: All related data loaded in ~3 queries instead of 500+
        invoices = query.order_by(DURPAInvoice.created_at.desc()).offset(skip).limit(limit).all()

        # Build response - NO additional queries needed
        records = []
        for inv in invoices:
            # Build items array - data already loaded
            items = []
            for item in inv.items:
                items.append({
                    "id": item.id,
                    "invoice_id": item.invoice_id,
                    "description_id": item.description_id,
                    "li_number": item.li_number,
                    "quantity": item.quantity,
                    "unit_price": item.unit_price,
                    "pac_date": item.pac_date,
                    "description_text": item.description.description if item.description else None
                })

            records.append({
                "id": inv.id,
                "project_id": inv.project_id,
                "po_number": inv.project.po_number,
                "ppo_number": inv.ppo_number,
                "new_po_number": inv.new_po_number,
                "pr_number": inv.pr_number,
                "site_id": inv.site_id,
                "model": inv.model,
                "sap_invoice_number": inv.sap_invoice_number,
                "invoice_date": inv.invoice_date,
                "customer_invoice_number": inv.customer_invoice_number,
                "prf_percentage": inv.prf_percentage,
                "vat_rate": inv.vat_rate,
                "created_at": inv.created_at,
                "items": items,
                "items_count": len(items)
            })

        return {"records": records, "total": total}

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error retrieving invoices: {str(e)}"
        )


@duRPALogisticsRoute.get("/du-rpa/invoices/{invoice_id}", response_model=DURPAInvoiceOut)
def get_invoice(
        invoice_id: int,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    """Get a specific invoice by ID. OPTIMIZED: Uses eager loading."""
    from sqlalchemy.orm import joinedload

    # OPTIMIZED: Eager load items and descriptions
    invoice = db.query(DURPAInvoice).filter(DURPAInvoice.id == invoice_id).options(
        joinedload(DURPAInvoice.items).joinedload(DURPAInvoiceItem.description)
    ).first()

    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")

    # Build response - NO additional queries needed
    items = []
    for item in invoice.items:
        items.append(DURPAInvoiceItemOut(
            id=item.id,
            invoice_id=item.invoice_id,
            description_id=item.description_id,
            li_number=item.li_number,
            quantity=item.quantity,
            unit_price=item.unit_price,
            pac_date=item.pac_date,
            description_text=item.description.description if item.description else None
        ))

    return DURPAInvoiceOut(
        id=invoice.id,
        project_id=invoice.project_id,
        ppo_number=invoice.ppo_number,
        new_po_number=invoice.new_po_number,
        pr_number=invoice.pr_number,
        site_id=invoice.site_id,
        model=invoice.model,
        sap_invoice_number=invoice.sap_invoice_number,
        invoice_date=invoice.invoice_date,
        customer_invoice_number=invoice.customer_invoice_number,
        prf_percentage=invoice.prf_percentage,
        vat_rate=invoice.vat_rate,
        created_at=invoice.created_at,
        items=items
    )


@duRPALogisticsRoute.patch("/du-rpa/invoices/{invoice_id}/customer-invoice")
async def update_customer_invoice(
        invoice_id: int,
        customer_invoice: str,
        request: Request,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    """Update customer invoice number for an invoice."""
    invoice = db.query(DURPAInvoice).filter(DURPAInvoice.id == invoice_id).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")

    try:
        old_value = invoice.customer_invoice_number
        invoice.customer_invoice_number = customer_invoice.strip() if customer_invoice else None
        db.commit()

        await create_audit_log(
            db=db,
            user_id=current_user.id,
            action="update_du_rpa_invoice_customer_invoice",
            resource_type="du_rpa_invoice",
            resource_id=str(invoice_id),
            resource_name=invoice.ppo_number,
            details=json.dumps({"old": old_value, "new": invoice.customer_invoice_number}),
            ip_address=get_client_ip(request),
            user_agent=request.headers.get("User-Agent")
        )

        return {"message": "Customer invoice updated successfully", "customer_invoice_number": invoice.customer_invoice_number}

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error updating customer invoice: {str(e)}"
        )


@duRPALogisticsRoute.delete("/du-rpa/invoices/{invoice_id}")
async def delete_invoice(
        invoice_id: int,
        request: Request,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    """Delete an invoice."""
    invoice = db.query(DURPAInvoice).filter(DURPAInvoice.id == invoice_id).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")

    try:
        invoice_num = invoice.sap_invoice_number or str(invoice.id)
        db.delete(invoice)
        db.commit()

        await create_audit_log(
            db=db,
            user_id=current_user.id,
            action="delete_du_rpa_invoice",
            resource_type="du_rpa_invoice",
            resource_id=str(invoice_id),
            resource_name=invoice_num,
            ip_address=get_client_ip(request),
            user_agent=request.headers.get("User-Agent")
        )

        return {"message": "Invoice deleted successfully"}

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error deleting invoice: {str(e)}"
        )


@duRPALogisticsRoute.delete("/du-rpa/projects/{project_id}/invoices")
async def delete_all_invoices(
        project_id: int,
        request: Request,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    """Delete all invoices for a project."""
    project = db.query(DURPAProject).filter(DURPAProject.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    try:
        count = db.query(DURPAInvoice).filter(DURPAInvoice.project_id == project_id).count()
        db.query(DURPAInvoice).filter(DURPAInvoice.project_id == project_id).delete(synchronize_session=False)
        db.commit()

        await create_audit_log(
            db=db,
            user_id=current_user.id,
            action="delete_all_du_rpa_invoices",
            resource_type="du_rpa_invoice",
            resource_id=str(project_id),
            resource_name=project.po_number,
            details=json.dumps({"deleted_count": count}),
            ip_address=get_client_ip(request),
            user_agent=request.headers.get("User-Agent")
        )

        return {"message": f"Deleted {count} invoices"}

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error deleting invoices: {str(e)}"
        )


# ===========================
# ONE-TIME BULK IMPORT
# ===========================

@duRPALogisticsRoute.post("/du-rpa/import-consolidated-tracker")
async def import_consolidated_tracker(
        file: UploadFile = File(...),
        request: Request = None,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    """
    One-time bulk import from a multi-sheet Excel file (.xlsb / .xlsx).

    Each sheet whose name matches an active PO# is processed:
      - Rows 1-4 (transposed): PO Line Item, PO Qty as per PO, PO Qty per unit, Price per Unit
      - Row 11: Description names (columns F+)
      - Row 12+: Invoice rows  (A=Site ID, B=SAP Invoice#, C=Invoice Date,
                                 D=Customer Invoice#, E=PPO#, F+=quantities per description)
      - Trailing columns: VAT %, PRF % (auto-detected by header in row 11)

    Creates Projects, Descriptions, Invoices, and Invoice Items.
    """
    ACTIVE_POS = {
        '81136', '86301', '88436', '88527', '94892', '78881', '78882',
        '76738', '92154', '76732', '75948', '98101', '84411', '83923',
        '106745', '76748', '104576', '76826'
    }

    fname = file.filename.lower()
    if not fname.endswith(('.xlsb', '.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File must be an Excel file (.xlsb or .xlsx)")

    try:
        content = await file.read()
        engine = 'pyxlsb' if fname.endswith('.xlsb') else None
        xls = pd.ExcelFile(BytesIO(content), engine=engine)

        # Only process sheets whose name matches an active PO
        sheets_to_process = [s for s in xls.sheet_names if str(s).strip() in ACTIVE_POS]
        if not sheets_to_process:
            raise HTTPException(status_code=400, detail="No sheets found matching active PO numbers")

        stats = {
            'projects_created': 0,
            'descriptions_created': 0,
            'invoices_created': 0,
            'items_created': 0,
            'errors': []
        }

        # Pre-fetch existing PPO#s to skip duplicates
        existing_ppos = {p[0] for p in db.query(DURPAInvoice.ppo_number).all()}

        # ---- Helpers ----
        def parse_float(val):
            """Safely parse a float, returning None for NaN/empty."""
            if pd.isna(val):
                return None
            try:
                return float(val)
            except (ValueError, TypeError):
                return None

        def parse_currency(val):
            """Parse a currency string like '$1,140,327.80' to float."""
            if pd.isna(val):
                return None
            try:
                s = str(val).replace('$', '').replace(',', '').strip()
                return float(s) if s and s != 'nan' else None
            except (ValueError, TypeError):
                return None

        def parse_date(val):
            """Parse a date from either a datetime, string, or Excel serial number."""
            if pd.isna(val):
                return None
            try:
                # pyxlsb returns dates as float serial numbers
                if isinstance(val, (int, float)):
                    serial = float(val)
                    if serial > 30000:  # Looks like an Excel serial date
                        return (datetime(1899, 12, 30) + timedelta(days=serial)).date()
                return pd.to_datetime(val).date()
            except Exception:
                return None

        def safe_str(val):
            """Convert to string, return None for NaN/empty."""
            if pd.isna(val):
                return None
            s = str(val).strip()
            return s if s and s != 'nan' else None

        # ---- Process each sheet ----
        for sheet_name in sheets_to_process:
            po_number = str(sheet_name).strip()

            try:
                raw = pd.read_excel(xls, sheet_name=sheet_name, header=None)
            except Exception as e:
                stats['errors'].append(f"PO {po_number}: Could not read sheet — {e}")
                continue

            if len(raw) < 12:
                stats['errors'].append(f"PO {po_number}: Sheet has fewer than 12 rows, skipping")
                continue

            # --- Build a column map from the header row (row 10) ---
            # Column layout varies between sheets, so detect dynamically
            col_map = {}  # e.g. {'site id': 0, 'sap invoice #': 1, ...}
            for col_idx in range(len(raw.columns)):
                header_val = safe_str(raw.iloc[10, col_idx])
                if header_val:
                    col_map[header_val.lower()] = col_idx

            # Find key column indices
            site_col = col_map.get('site id')
            sap_col = col_map.get('sap invoice #') or col_map.get('sap invoice')
            date_col = col_map.get('invoice date')
            cust_col = col_map.get('customer invoice #') or col_map.get('customer invoice')
            ppo_col = col_map.get('po release #') or col_map.get('ppo#') or col_map.get('ppo')
            vat_col = col_map.get('vat %') or col_map.get('vat%') or col_map.get('vat rate')
            prf_col = col_map.get('prf %') or col_map.get('prf%') or col_map.get('prf percentage')

            # Determine category: PPO-based if sheet has a PO Release / PPO column
            category = 'ppo_based' if ppo_col is not None else 'non_ppo'

            # --- Project ---
            project = db.query(DURPAProject).filter(DURPAProject.po_number == po_number).first()
            if project:
                stats['errors'].append(f"PO {po_number}: Project already exists, adding data to it")
                # Update category if it changed
                if project.category != category:
                    project.category = category
            else:
                project = DURPAProject(po_number=po_number, category=category)
                db.add(project)
                db.flush()
                stats['projects_created'] += 1

            # --- Identify description columns (col 5+ where row 0 has a PO Line Item) ---
            desc_columns = []  # list of (col_idx, desc_text)
            for col_idx in range(5, len(raw.columns)):
                li_val = raw.iloc[0, col_idx]
                if pd.isna(li_val):
                    continue
                desc_text = safe_str(raw.iloc[10, col_idx])
                if not desc_text:
                    continue
                desc_columns.append((col_idx, desc_text))

            if not desc_columns:
                stats['errors'].append(f"PO {po_number}: No description columns found, skipping")
                continue

            # --- Create descriptions ---
            existing_descs = db.query(DURPADescription).filter(
                DURPADescription.project_id == project.id
            ).all()
            desc_map = {d.description.strip().lower(): d for d in existing_descs}

            for col_idx, desc_text in desc_columns:
                key = desc_text.lower()
                if key in desc_map:
                    continue

                po_line_item = safe_str(raw.iloc[0, col_idx])
                po_qty_as_per_po = parse_float(raw.iloc[1, col_idx])
                po_qty_per_unit = parse_float(raw.iloc[2, col_idx])
                price_per_unit = parse_currency(raw.iloc[3, col_idx])

                desc = DURPADescription(
                    project_id=project.id,
                    description=desc_text,
                    po_line_item=po_line_item,
                    po_qty_as_per_po=po_qty_as_per_po,
                    po_qty_per_unit=po_qty_per_unit,
                    price_per_unit=price_per_unit
                )
                db.add(desc)
                db.flush()
                desc_map[key] = desc
                stats['descriptions_created'] += 1

            # --- Parse invoice rows (row 11 onwards, 0-indexed) ---
            for row_idx in range(11, len(raw)):
                # Determine PPO#: use dedicated column if it exists,
                # otherwise construct from PO# + SAP Invoice # or Customer Invoice #
                if ppo_col is not None:
                    ppo_number = safe_str(raw.iloc[row_idx, ppo_col])
                else:
                    # No PPO column — build a unique key from available data
                    sap_val = safe_str(raw.iloc[row_idx, sap_col]) if sap_col is not None else None
                    cust_val = safe_str(raw.iloc[row_idx, cust_col]) if cust_col is not None else None
                    identifier = sap_val or cust_val
                    ppo_number = f"{po_number}-{identifier}" if identifier else None

                if not ppo_number:
                    continue  # Empty row

                # Check if row has any description quantities (skip truly empty rows)
                has_data = any(
                    parse_float(raw.iloc[row_idx, ci]) not in (None, 0)
                    for ci, _ in desc_columns
                )
                if not has_data:
                    continue

                if ppo_number in existing_ppos:
                    stats['errors'].append(f"PPO# {ppo_number}: Already exists, skipping")
                    continue

                # Parse VAT %
                vat_val = None
                if vat_col is not None:
                    vat_val = parse_float(raw.iloc[row_idx, vat_col])
                    if vat_val is not None and vat_val < 1:
                        vat_val = round(vat_val * 100, 2)

                # Parse PRF %
                prf_val = None
                if prf_col is not None:
                    prf_val = parse_float(raw.iloc[row_idx, prf_col])
                    if prf_val is not None and prf_val < 1:
                        prf_val = round(prf_val * 100, 2)

                invoice = DURPAInvoice(
                    project_id=project.id,
                    ppo_number=ppo_number,
                    new_po_number=po_number,
                    site_id=safe_str(raw.iloc[row_idx, site_col]) if site_col is not None else None,
                    sap_invoice_number=safe_str(raw.iloc[row_idx, sap_col]) if sap_col is not None else None,
                    invoice_date=parse_date(raw.iloc[row_idx, date_col]) if date_col is not None else None,
                    customer_invoice_number=safe_str(raw.iloc[row_idx, cust_col]) if cust_col is not None else None,
                    vat_rate=vat_val,
                    prf_percentage=prf_val
                )
                db.add(invoice)
                db.flush()
                stats['invoices_created'] += 1
                existing_ppos.add(ppo_number)

                # --- Create invoice items for each description column ---
                for col_idx, desc_text in desc_columns:
                    qty = parse_float(raw.iloc[row_idx, col_idx])
                    if qty is None or qty == 0:
                        continue

                    desc = desc_map.get(desc_text.lower())
                    if not desc:
                        continue

                    item = DURPAInvoiceItem(
                        invoice_id=invoice.id,
                        description_id=desc.id,
                        li_number=safe_str(raw.iloc[0, col_idx]),
                        quantity=qty,
                        unit_price=desc.price_per_unit
                    )
                    db.add(item)
                    stats['items_created'] += 1

        db.commit()

        await create_audit_log(
            db=db,
            user_id=current_user.id,
            action="import_consolidated_tracker",
            resource_type="du_rpa_project",
            resource_id=None,
            resource_name=file.filename,
            details=json.dumps({
                "projects": stats['projects_created'],
                "descriptions": stats['descriptions_created'],
                "invoices": stats['invoices_created'],
                "items": stats['items_created'],
                "errors_count": len(stats['errors'])
            }),
            ip_address=get_client_ip(request),
            user_agent=request.headers.get("User-Agent")
        )

        msg = (
            f"Import complete: {stats['projects_created']} projects, "
            f"{stats['descriptions_created']} descriptions, "
            f"{stats['invoices_created']} invoices, "
            f"{stats['items_created']} items created"
        )
        return UploadResponse(inserted=stats['invoices_created'], errors=stats['errors'], message=msg)

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Error importing tracker: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error importing data: {str(e)}"
        )


def _generate_excel(invoice, vat_rate, exchange_rate, template_path):
    """Generate the Excel invoice from template. Returns BytesIO."""
    wb = openpyxl.load_workbook(template_path)
    ws = wb['Invoice']

    num_items = len(invoice.items)
    MAX_ITEMS_NO_SHIFT = 34
    shift = max(0, num_items - MAX_ITEMS_NO_SHIFT)

    if shift > 0:
        ws.insert_rows(47, shift)
        r49 = 49 + shift
        r51 = 51 + shift
        r52 = 52 + shift
        r54 = 54 + shift
        r56 = 56 + shift
        r57 = 57 + shift
        r69 = 69 + shift
        r71 = 71 + shift
        r78 = 78 + shift
        data_end = 68 + shift

        ws.cell(row=r49, column=3).value = f'=-$G${r69}*0'
        ws.cell(row=r49, column=4).value = '=0*0.05'
        ws.cell(row=r49, column=5).value = f'=SUM($C${r49}:$D${r49})'
        ws.cell(row=r51, column=3).value = f'=$G${r69}+$C${r49}'
        ws.cell(row=r51, column=4).value = f'=($G${r69}*0.05)+$D${r49}'
        ws.cell(row=r51, column=5).value = f'=SUM($C${r51}:$D${r51})'
        ws.cell(row=r52, column=4).value = f'=$D${r51}*$G${r78}'
        ws.cell(row=r54, column=5).value = f'=SUM($E${r54 + 1}:$E${r57})'
        ws.cell(row=r56, column=5).value = f'=$G${r69}+$C${r49}-$E${r57}+$D${r51}'
        ws.cell(row=r57, column=5).value = f'=$G${r69}*0.1'
        ws.cell(row=r69, column=7).value = f'=SUM(G13:G{data_end})'
        ws.cell(row=r69, column=9).value = f'=SUM(I13:I{data_end})'
        ws.cell(row=r69, column=10).value = f'=SUM(J13:J{data_end})'
        ws.cell(row=r71, column=7).value = f'=G{r69}*$G${r78}'
        ws.cell(row=r71, column=9).value = f'=I{r69}*$G${r78}'
        ws.cell(row=r71, column=10).value = f'=J{r69}*$G${r78}'
        ws.cell(row=r78, column=7).value = exchange_rate
    else:
        ws['G78'] = exchange_rate

    ws['E3'] = invoice.customer_invoice_number or ''
    ws['E4'] = invoice.ppo_number or ''
    if invoice.invoice_date:
        ws['G4'] = invoice.invoice_date

    pac_date = None
    for item in invoice.items:
        if item.pac_date:
            pac_date = item.pac_date
            break
    if pac_date:
        ws['O7'] = pac_date

    ws['O5'] = vat_rate
    ws['O6'] = exchange_rate

    thin_bottom = openpyxl.styles.Side(style='thin')

    for idx, item in enumerate(invoice.items):
        row = 13 + idx
        qty = item.quantity or 0
        unit_price = item.unit_price or (item.description.price_per_unit if item.description else 0) or 0
        amount = qty * unit_price
        vat_amount = vat_rate * amount

        ws.cell(row=row, column=1, value=invoice.site_id or '')
        ws.cell(row=row, column=2, value=item.description.description if item.description else '')
        ws.cell(row=row, column=3, value=item.li_number or '')
        ws.cell(row=row, column=4, value=f'=IF(ISBLANK(C{row}),"",IF($O$2="FIDX","Services","Equipment"))')
        ws.cell(row=row, column=5, value=qty)
        ws.cell(row=row, column=6, value=unit_price)
        ws.cell(row=row, column=7, value=amount)
        ws.cell(row=row, column=8, value=vat_rate)
        ws.cell(row=row, column=9, value=vat_amount)
        ws.cell(row=row, column=10, value=amount + vat_amount)

        # Bottom border on every item row
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col)
            cur = cell.border
            cell.border = openpyxl.styles.Border(
                left=cur.left, right=cur.right, top=cur.top,
                bottom=thin_bottom
            )

    for sheet_name in list(wb.sheetnames):
        if sheet_name != 'Invoice':
            del wb[sheet_name]

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _excel_to_pdf(excel_data):
    """Convert Excel bytes to PDF using Excel's built-in PDF export via win32com.

    This produces a pixel-perfect PDF since Excel renders directly from the
    template with all its formatting, images, and page layout intact.
    """
    import tempfile
    import win32com.client

    # win32com needs real file paths on disk
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        tmp.write(excel_data)
        xlsx_path = Path(tmp.name).resolve()

    pdf_path = xlsx_path.with_suffix('.pdf')
    app = None
    wb = None

    try:
        app = win32com.client.Dispatch("Excel.Application")
        app.Visible = False
        app.DisplayAlerts = False

        wb = app.Workbooks.Open(str(xlsx_path))
        wb.ExportAsFixedFormat(0, str(pdf_path))   # 0 = xlTypePDF

        return BytesIO(pdf_path.read_bytes())
    finally:
        # Always close workbook and quit Excel, even on error
        if wb:
            try:
                wb.Close(False)
            except Exception:
                pass
        if app:
            try:
                app.Quit()
            except Exception:
                pass
        # Always clean up temp files
        xlsx_path.unlink(missing_ok=True)
        pdf_path.unlink(missing_ok=True)


@duRPALogisticsRoute.get("/du-rpa/invoices/{invoice_id}/download-excel")
async def download_invoice_excel(
        invoice_id: int,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    """
    Download an invoice as a ZIP containing both an Excel (.xlsx) and PDF file.

    The Excel is generated from the DU Invoice Template.
    The PDF is generated via Excel's own PDF export (win32com) for pixel-perfect output.
    Handles large invoices (100+ items) by shifting footer sections down in Excel.
    """
    from utils.exchange_rate import get_usd_aed_rate

    # Fetch invoice with items and descriptions
    invoice = db.query(DURPAInvoice).filter(DURPAInvoice.id == invoice_id).options(
        joinedload(DURPAInvoice.items).joinedload(DURPAInvoiceItem.description)
    ).first()

    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")

    # Get template path
    template_path = Path(__file__).parent.parent.parent / "templates" / "DU Invoice Template.xlsx"
    if not template_path.exists():
        raise HTTPException(status_code=500, detail="Invoice template not found")

    # Fetch live exchange rate
    try:
        rate_data = await get_usd_aed_rate()
        exchange_rate = rate_data["rate"]
    except Exception:
        exchange_rate = 3.6725  # Fallback

    vat_rate = (invoice.vat_rate or 5) / 100

    try:
        raw_name = f"Invoice_{invoice.customer_invoice_number or invoice.ppo_number or invoice.id}"
        base_name = "".join(c if (c.isalnum() or c in '-_.') else '_' for c in raw_name)

        # Generate Excel, then convert it to PDF via Excel's own renderer
        excel_buf = _generate_excel(invoice, vat_rate, exchange_rate, template_path)
        excel_data = excel_buf.read()
        pdf_buf = _excel_to_pdf(excel_data)

        # Create ZIP
        zip_output = BytesIO()
        with zipfile.ZipFile(zip_output, 'w', zipfile.ZIP_DEFLATED) as zf:
            zf.writestr(f"{base_name}.xlsx", excel_data)
            zf.writestr(f"{base_name}.pdf", pdf_buf.read())
        zip_output.seek(0)

        return StreamingResponse(
            zip_output,
            media_type="application/zip",
            headers={
                "Content-Disposition": f"attachment; filename={base_name}.zip"
            }
        )

    except Exception as e:
        logger.error(f"Error generating invoice download: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error generating invoice download: {str(e)}"
        )
