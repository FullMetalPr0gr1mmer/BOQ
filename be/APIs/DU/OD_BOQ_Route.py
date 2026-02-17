"""
OD BOQ Route - APIs for the new 3-table structure

Tables:
1. OD_BOQ_Site: Parent table (sites/projects)
2. OD_BOQ_Product: Product catalog
3. OD_BOQ_Site_Product: Junction table with quantities

Features:
- Full CRUD for sites, products, and site-products
- CSV upload with automatic population of all 3 tables
- Pagination and filtering
- Admin control and audit logging
- Site with products retrieval (join)
"""

import json
import csv
import logging
import os
import pandas as pd
from io import StringIO
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query, status, Request, Form
from sqlalchemy.orm import Session
from sqlalchemy import or_, func
from typing import Optional, List, Dict, Any, Union, Tuple

# Configure logging
logger = logging.getLogger(__name__)

# ===========================
# MODULE CONSTANTS
# ===========================

# Pagination defaults
DEFAULT_PAGINATION_SKIP = 0
DEFAULT_SITES_LIMIT = 50
DEFAULT_PRODUCTS_LIMIT = 100
MAX_SITES_LIMIT = 500
MAX_PRODUCTS_LIMIT = 1000

# CSV structure constants
CSV_PRODUCT_START_COL = 7  # Column index where product quantities begin
CSV_METADATA_FIELD_COUNT = 13  # Number of metadata fields per site row

# Date/time formats
DATE_FORMAT_DISPLAY = '%d-%b-%y'  # e.g., "09-Feb-26"
DATE_FORMAT_ISO = '%Y-%m-%d'  # e.g., "2026-02-09"

# Excel template path
BOQ_TEMPLATE_FILENAME = 'BOQ Formate.xlsx'
BOQ_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), '..', '..', 'templates', BOQ_TEMPLATE_FILENAME)

# Excel column mapping for BOQ data
BOQ_EXCEL_COLUMNS = [
    (1, ''),           # Empty column A
    (2, 'line'),       # Line number
    (3, 'bu'),         # Business Unit
    (4, 'item_job'),   # Item/Job code
    (5, 'description'),# Description
    (6, ''),           # Empty column F
    (7, 'qty'),        # Quantity
    (8, 'unit_price'), # Unit price
    (9, 'total_usd'),  # Total USD
    (10, 'total_aed'), # Total AED
    (11, 'site_id_list'), # Site ID
    (12, 'po_model')   # Model Name towards dU
]

from APIs.Core import get_db, get_current_user
from Models.DU.OD_BOQ_Site import ODBOQSite
from Models.DU.OD_BOQ_Product import ODBOQProduct
from Models.DU.OD_BOQ_Site_Product import ODBOQSiteProduct
import importlib
DU_Project_module = importlib.import_module("Models.DU.DU_Project")
DUProject = DU_Project_module.DUProject

from Models.Admin.User import User, UserProjectAccess
from Models.Admin.AuditLog import AuditLog
from Schemas.DU.OD_BOQ_Schema import (
    ODBOQSiteCreate, ODBOQSiteUpdate, ODBOQSiteOut, ODBOQSitePagination,
    ODBOQProductCreate, ODBOQProductUpdate, ODBOQProductOut, ODBOQProductPagination,
    ODBOQSiteProductCreate, ODBOQSiteProductUpdate, ODBOQSiteProductOut,
    SiteWithProductsOut, ProductWithQuantity,
    ODBOQStatsResponse, FilterOptions, UploadResponse, BulkDeleteResponse
)

odBOQRoute = APIRouter(prefix="/od-boq", tags=["OD BOQ (New Structure)"])


# ===========================
# HELPER FUNCTIONS
# ===========================

async def create_audit_log(
        db: Session,
        user_id: int,
        action: str,
        resource_type: str,
        resource_id: Optional[str] = None,
        resource_name: Optional[str] = None,
        details: Optional[str] = None,
        ip_address: Optional[str] = None,
        user_agent: Optional[str] = None
):
    """Create an audit log entry."""
    audit_log = AuditLog(
        user_id=user_id,
        action=action,
        resource_type=resource_type,
        resource_id=resource_id,
        resource_name=resource_name,
        details=details,
        ip_address=ip_address,
        user_agent=user_agent
    )
    db.add(audit_log)
    db.commit()
    return audit_log


def get_client_ip(request: Request) -> str:
    """Extract client IP from request."""
    forwarded = request.headers.get("X-Forwarded-For")
    if forwarded:
        return forwarded.split(",")[0].strip()
    return request.client.host if request.client else "unknown"


def check_du_project_access(current_user: User, project: DUProject, db: Session, required_permission: str = "view"):
    """
    Helper function to check if user has access to a DU project with required permission level.
    """
    # Senior admin has all permissions to all projects
    if current_user.role.name == "senior_admin":
        return True

    # Admin has all permissions but only to projects they have access to
    # Check UserProjectAccess for both admin and other roles
    access = db.query(UserProjectAccess).filter(
        UserProjectAccess.user_id == current_user.id,
        UserProjectAccess.DUproject_id == project.pid_po
    ).first()

    if not access:
        return False

    # Admin with any access level gets full permissions (same as senior_admin) for their projects
    if current_user.role.name == "admin":
        return True

    # For non-admin users, check permission levels
    permission_hierarchy = {
        "view": ["view", "edit", "all"],
        "edit": ["edit", "all"],
        "all": ["all"]
    }

    return access.permission_level in permission_hierarchy.get(required_permission, [])


def validate_site_project_access(
    site: ODBOQSite,
    current_user: User,
    db: Session,
    required_permission: str = "view"
) -> Tuple[bool, Optional[str]]:
    """
    Validate user access to a site's project.

    Returns:
        Tuple of (has_access: bool, error_message: Optional[str])
    """
    if not site.project_id:
        return True, None

    project = db.query(DUProject).filter(DUProject.pid_po == site.project_id).first()
    if not project:
        return True, None  # No project means no restriction

    if not check_du_project_access(current_user, project, db, required_permission):
        return False, f"You do not have {required_permission} access to this site's project."

    return True, None


def clean_csv_value(value) -> Optional[str]:
    """Extract and clean a CSV value, returning None if empty."""
    if pd.isna(value):
        return None
    val = str(value).strip()
    return val if val else None


def safe_extract_float(value) -> Optional[float]:
    """Safely convert a value to float for quantities."""
    if pd.isna(value):
        return None
    val_str = str(value).strip()
    if not val_str:
        return None
    try:
        return float(val_str)
    except ValueError:
        return None


def parse_currency_value(value) -> Optional[float]:
    """Parse a currency string like ' $1,126.28 ' to float."""
    if pd.isna(value):
        return None
    val_str = str(value).strip()
    if not val_str:
        return None
    # Remove $, commas, and whitespace
    val_str = val_str.replace('$', '').replace(',', '').strip()
    if not val_str:
        return None
    try:
        return float(val_str)
    except ValueError:
        return None


def format_quantity(qty_value: float) -> Union[int, float]:
    """Convert float to int if it's a whole number (e.g., 2.0 -> 2)."""
    if isinstance(qty_value, float) and qty_value == int(qty_value):
        return int(qty_value)
    return qty_value


def write_boq_row_to_excel(ws, row_num: int, item: Dict[str, Any], template_border) -> None:
    """Write a single BOQ item row to the worksheet with borders."""
    for col_idx, field in BOQ_EXCEL_COLUMNS:
        value = item.get(field) if field else ''
        ws.cell(row=row_num, column=col_idx).value = value
        ws.cell(row=row_num, column=col_idx).border = template_border


def get_user_accessible_du_projects(current_user: User, db: Session):
    """Get all DU projects that the current user has access to."""
    # Senior admin can see all projects
    if current_user.role.name == "senior_admin":
        return db.query(DUProject).all()

    # For admin and other users, get projects they have access to
    user_accesses = db.query(UserProjectAccess).filter(
        UserProjectAccess.user_id == current_user.id,
        UserProjectAccess.DUproject_id.isnot(None)
    ).all()

    if not user_accesses:
        return []

    accessible_project_ids = [access.DUproject_id for access in user_accesses]
    return db.query(DUProject).filter(DUProject.pid_po.in_(accessible_project_ids)).all()


def filter_sites_by_user_access(current_user: User, query, db: Session):
    """Filter sites query based on user's project access."""
    if current_user.role.name == "senior_admin":
        return query

    accessible_projects = get_user_accessible_du_projects(current_user, db)
    accessible_project_ids = [project.pid_po for project in accessible_projects]

    if not accessible_project_ids:
        return query.filter(ODBOQSite.site_id == "___NONE___")  # Return empty

    return query.filter(ODBOQSite.project_id.in_(accessible_project_ids))


# ===========================
# SITE CRUD ENDPOINTS
# ===========================

@odBOQRoute.post("/sites", response_model=ODBOQSiteOut)
async def create_site(
        site_data: ODBOQSiteCreate,
        request: Request,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    """Create a new site."""
    # Check project access if project_id is provided
    if site_data.project_id:
        project = db.query(DUProject).filter(DUProject.pid_po == site_data.project_id).first()
        if not project:
            raise HTTPException(status_code=404, detail="Project not found")

        if not check_du_project_access(current_user, project, db, "edit"):
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You are not authorized to add sites to this project."
            )

    # Check if site_id with same subscope already exists (unique constraint)
    existing_site = db.query(ODBOQSite).filter(
        ODBOQSite.site_id == site_data.site_id,
        ODBOQSite.subscope == site_data.subscope
    ).first()
    if existing_site:
        raise HTTPException(status_code=400, detail=f"Site with ID '{site_data.site_id}' and subscope '{site_data.subscope}' already exists")

    try:
        new_site = ODBOQSite(**site_data.dict())
        db.add(new_site)
        db.commit()
        db.refresh(new_site)

        await create_audit_log(
            db=db,
            user_id=current_user.id,
            action="create_od_boq_site",
            resource_type="od_boq_site",
            resource_id=new_site.site_id,
            resource_name=new_site.site_id,
            details=json.dumps({"project_id": site_data.project_id}),
            ip_address=get_client_ip(request),
            user_agent=request.headers.get("User-Agent")
        )

        return new_site

    except Exception as e:
        db.rollback()
        logger.error(f"Error creating site: {str(e)}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="An error occurred while creating the site. Please try again."
        )


@odBOQRoute.get("/sites", response_model=ODBOQSitePagination)
def get_sites(
        skip: int = Query(0, ge=0),
        limit: int = Query(50, ge=1, le=500),
        search: Optional[str] = None,
        project_id: Optional[str] = None,
        scope: Optional[str] = None,
        subscope: Optional[str] = None,
        region: Optional[str] = None,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    """Get sites with pagination and filters."""
    try:
        query = db.query(ODBOQSite)
        query = filter_sites_by_user_access(current_user, query, db)

        # Apply filters
        if project_id:
            query = query.filter(ODBOQSite.project_id == project_id)
        if scope:
            query = query.filter(ODBOQSite.scope == scope)
        if subscope:
            query = query.filter(ODBOQSite.subscope == subscope)
        if region:
            query = query.filter(ODBOQSite.region == region)

        # Search
        if search:
            search_pattern = f"%{search}%"
            query = query.filter(
                or_(
                    ODBOQSite.site_id.ilike(search_pattern),
                    ODBOQSite.scope.ilike(search_pattern),
                    ODBOQSite.subscope.ilike(search_pattern),
                    ODBOQSite.po_model.ilike(search_pattern)
                )
            )

        total_count = query.count()
        records = query.order_by(ODBOQSite.site_id).offset(skip).limit(limit).all()

        return ODBOQSitePagination(records=records, total=total_count)

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error retrieving sites: {str(e)}"
        )


@odBOQRoute.get("/sites/{id}", response_model=ODBOQSiteOut)
def get_site(
        id: int,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    """Get a specific site by database ID."""
    site = db.query(ODBOQSite).filter(ODBOQSite.id == id).first()
    if not site:
        raise HTTPException(status_code=404, detail="Site not found")

    # Check access
    if site.project_id:
        project = db.query(DUProject).filter(DUProject.pid_po == site.project_id).first()
        if project and not check_du_project_access(current_user, project, db, "view"):
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You are not authorized to view this site."
            )

    return site


@odBOQRoute.get("/sites/{id}/with-products", response_model=SiteWithProductsOut)
def get_site_with_products(
        id: int,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    """Get a site with all its products and quantities."""
    site = db.query(ODBOQSite).filter(ODBOQSite.id == id).first()
    if not site:
        raise HTTPException(status_code=404, detail="Site not found")

    # Check access
    if site.project_id:
        project = db.query(DUProject).filter(DUProject.pid_po == site.project_id).first()
        if project and not check_du_project_access(current_user, project, db, "view"):
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You are not authorized to view this site."
            )

    # Get all site-product records for this site with product details
    site_products = db.query(
        ODBOQSiteProduct,
        ODBOQProduct
    ).join(
        ODBOQProduct, ODBOQSiteProduct.product_id == ODBOQProduct.id
    ).filter(
        ODBOQSiteProduct.site_record_id == site.id
    ).all()

    products_list = []
    total_qty_sum = 0.0
    for sp, product in site_products:
        products_list.append({
            "product_id": product.id,
            "description": product.description,
            "line_number": product.line_number,
            "code": product.code,
            "category": product.category,
            "unit_price": product.unit_price,
            "total_po_qty": product.total_po_qty,
            "consumed_in_year": product.consumed_in_year,
            "consumed_year": product.consumed_year,
            "remaining_in_po": product.remaining_in_po,
            "qty_per_site": sp.qty_per_site
        })
        if sp.qty_per_site:
            total_qty_sum += sp.qty_per_site

    return SiteWithProductsOut(
        id=site.id,
        site_id=site.site_id,
        region=site.region,
        distance=site.distance,
        scope=site.scope,
        subscope=site.subscope,
        po_model=site.po_model,
        project_id=site.project_id,
        ac_armod_cable=site.ac_armod_cable,
        additional_cost=site.additional_cost,
        remark=site.remark,
        partner=site.partner,
        request_status=site.request_status,
        requested_date=site.requested_date,
        du_po_number=site.du_po_number,
        smp=site.smp,
        year_scope=site.year_scope,
        integration_status=site.integration_status,
        integration_date=site.integration_date,
        du_po_convention_name=site.du_po_convention_name,
        po_year_issuance=site.po_year_issuance,
        products=products_list,
        total_qty_sum=total_qty_sum
    )


@odBOQRoute.put("/sites/{id}", response_model=ODBOQSiteOut)
async def update_site(
        id: int,
        site_data: ODBOQSiteUpdate,
        request: Request,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    """Update a site."""
    site = db.query(ODBOQSite).filter(ODBOQSite.id == id).first()
    if not site:
        raise HTTPException(status_code=404, detail="Site not found")

    # Check access
    if site.project_id:
        project = db.query(DUProject).filter(DUProject.pid_po == site.project_id).first()
        if project and not check_du_project_access(current_user, project, db, "edit"):
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You are not authorized to update this site."
            )

    try:
        update_data = site_data.dict(exclude_unset=True)
        for field, value in update_data.items():
            if value is not None:
                setattr(site, field, value)

        db.commit()
        db.refresh(site)

        await create_audit_log(
            db=db,
            user_id=current_user.id,
            action="update_od_boq_site",
            resource_type="od_boq_site",
            resource_id=str(site.id),
            resource_name=f"{site.site_id} ({site.subscope})",
            details=json.dumps(update_data),
            ip_address=get_client_ip(request),
            user_agent=request.headers.get("User-Agent")
        )

        return site

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error updating site: {str(e)}"
        )


@odBOQRoute.delete("/sites/{id}")
async def delete_site(
        id: int,
        request: Request,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    """Delete a site and all its site-product records."""
    site = db.query(ODBOQSite).filter(ODBOQSite.id == id).first()
    if not site:
        raise HTTPException(status_code=404, detail="Site not found")

    # Check access
    if site.project_id:
        project = db.query(DUProject).filter(DUProject.pid_po == site.project_id).first()
        if project and not check_du_project_access(current_user, project, db, "all"):
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You are not authorized to delete this site."
            )

    try:
        project_id = site.project_id
        site_identifier = f"{site.site_id} ({site.subscope})"

        # Delete all site-product records first
        db.query(ODBOQSiteProduct).filter(ODBOQSiteProduct.site_record_id == site.id).delete(synchronize_session=False)

        # Delete the site
        db.delete(site)
        db.commit()

        await create_audit_log(
            db=db,
            user_id=current_user.id,
            action="delete_od_boq_site",
            resource_type="od_boq_site",
            resource_id=str(id),
            resource_name=site_identifier,
            details=json.dumps({"project_id": project_id}),
            ip_address=get_client_ip(request),
            user_agent=request.headers.get("User-Agent")
        )

        return {"message": f"Site {site_identifier} and all its products deleted successfully"}

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error deleting site: {str(e)}"
        )


@odBOQRoute.delete("/sites/delete-all/{project_id}", response_model=BulkDeleteResponse)
async def delete_all_sites_by_project(
        project_id: str,
        request: Request,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    """Delete all sites and their site-product records for a specific project."""
    # Check if project exists and user has access
    project = db.query(DUProject).filter(DUProject.pid_po == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    if not check_du_project_access(current_user, project, db, "all"):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You are not authorized to delete sites from this project."
        )

    try:
        # Get all sites for this project
        sites = db.query(ODBOQSite).filter(ODBOQSite.project_id == project_id).all()
        site_record_ids = [site.id for site in sites]

        # Count records before deletion
        deleted_sites = len(site_record_ids)
        deleted_site_products = 0

        if site_record_ids:
            # Delete all site-product records for these sites
            deleted_site_products = db.query(ODBOQSiteProduct).filter(
                ODBOQSiteProduct.site_record_id.in_(site_record_ids)
            ).delete(synchronize_session=False)

            # Delete all sites
            db.query(ODBOQSite).filter(ODBOQSite.project_id == project_id).delete(synchronize_session=False)

        db.commit()

        await create_audit_log(
            db=db,
            user_id=current_user.id,
            action="bulk_delete_od_boq_sites",
            resource_type="od_boq_site",
            resource_id=project_id,
            resource_name=project_id,
            details=json.dumps({
                "project_id": project_id,
                "deleted_sites": deleted_sites,
                "deleted_site_products": deleted_site_products
            }),
            ip_address=get_client_ip(request),
            user_agent=request.headers.get("User-Agent")
        )

        return BulkDeleteResponse(
            deleted_sites=deleted_sites,
            deleted_site_products=deleted_site_products,
            message=f"Successfully deleted {deleted_sites} sites and {deleted_site_products} site-product records for project {project_id}"
        )

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error deleting sites: {str(e)}"
        )


# ===========================
# PRODUCT CRUD ENDPOINTS
# ===========================

@odBOQRoute.post("/products", response_model=ODBOQProductOut)
async def create_product(
        product_data: ODBOQProductCreate,
        request: Request,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    """Create a new product."""
    try:
        new_product = ODBOQProduct(**product_data.dict())
        db.add(new_product)
        db.commit()
        db.refresh(new_product)

        await create_audit_log(
            db=db,
            user_id=current_user.id,
            action="create_od_boq_product",
            resource_type="od_boq_product",
            resource_id=str(new_product.id),
            resource_name=new_product.description[:100] if new_product.description else None,
            ip_address=get_client_ip(request),
            user_agent=request.headers.get("User-Agent")
        )

        return new_product

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error creating product: {str(e)}"
        )


@odBOQRoute.get("/products", response_model=ODBOQProductPagination)
def get_products(
        skip: int = Query(0, ge=0),
        limit: int = Query(100, ge=1, le=1000),
        search: Optional[str] = None,
        category: Optional[str] = None,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    """Get products with pagination and filters."""
    try:
        query = db.query(ODBOQProduct)

        # Apply filters
        if category:
            query = query.filter(ODBOQProduct.category == category)

        # Search
        if search:
            search_pattern = f"%{search}%"
            query = query.filter(
                or_(
                    ODBOQProduct.description.ilike(search_pattern),
                    ODBOQProduct.code.ilike(search_pattern),
                    ODBOQProduct.line_number.ilike(search_pattern)
                )
            )

        total_count = query.count()
        records = query.order_by(ODBOQProduct.id).offset(skip).limit(limit).all()

        return ODBOQProductPagination(records=records, total=total_count)

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error retrieving products: {str(e)}"
        )


# ===========================
# STATISTICS & FILTER ENDPOINTS
# ===========================

@odBOQRoute.get("/stats", response_model=ODBOQStatsResponse)
def get_stats(
        project_id: Optional[str] = None,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    """Get overall statistics."""
    try:
        site_query = db.query(ODBOQSite)
        site_query = filter_sites_by_user_access(current_user, site_query, db)

        if project_id:
            site_query = site_query.filter(ODBOQSite.project_id == project_id)

        total_sites = site_query.count()
        total_products = db.query(ODBOQProduct).count()
        total_site_products = db.query(ODBOQSiteProduct).count()

        unique_scopes = site_query.with_entities(func.count(func.distinct(ODBOQSite.scope))).scalar() or 0
        unique_subscopes = site_query.with_entities(func.count(func.distinct(ODBOQSite.subscope))).scalar() or 0
        unique_categories = db.query(ODBOQProduct).with_entities(func.count(func.distinct(ODBOQProduct.category))).scalar() or 0

        return ODBOQStatsResponse(
            total_sites=total_sites,
            total_products=total_products,
            total_site_products=total_site_products,
            unique_scopes=unique_scopes,
            unique_subscopes=unique_subscopes,
            unique_categories=unique_categories
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error retrieving stats: {str(e)}"
        )


@odBOQRoute.get("/filters/options", response_model=FilterOptions)
def get_filter_options(
        project_id: Optional[str] = None,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    """Get available filter options."""
    try:
        site_query = db.query(ODBOQSite)
        site_query = filter_sites_by_user_access(current_user, site_query, db)

        if project_id:
            site_query = site_query.filter(ODBOQSite.project_id == project_id)

        regions = [r[0] for r in site_query.with_entities(ODBOQSite.region).distinct().all() if r[0]]
        scopes = [r[0] for r in site_query.with_entities(ODBOQSite.scope).distinct().all() if r[0]]
        subscopes = [r[0] for r in site_query.with_entities(ODBOQSite.subscope).distinct().all() if r[0]]
        categories = [r[0] for r in db.query(ODBOQProduct).with_entities(ODBOQProduct.category).distinct().all() if r[0]]
        projects = [r[0] for r in site_query.with_entities(ODBOQSite.project_id).distinct().all() if r[0]]

        return FilterOptions(
            regions=sorted(regions),
            scopes=sorted(scopes),
            subscopes=sorted(subscopes),
            categories=sorted(categories),
            projects=sorted(projects)
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error retrieving filter options: {str(e)}"
        )


# ===========================
# CSV UPLOAD ENDPOINT (Populates all 3 tables)
# ===========================

@odBOQRoute.post("/upload-csv", response_model=UploadResponse)
async def upload_csv(
        file: UploadFile = File(...),
        project_id: str = Form(...),
        consumed_year: int = Form(2026),
        request: Request = None,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    """
    Upload CSV file and populate all 3 tables:
    - Sites (rows 8+, columns A-F + metadata columns)
    - Products (rows 1-5, columns 7 onwards) - consumed_in_year and remaining_in_po auto-calculated
    - Site-Products (junction with quantities)

    Changes:
    - consumed_in_year: Auto-calculated as SUM of all site quantities for each product
    - remaining_in_po: Auto-calculated as total_po_qty - consumed_in_year
    - Same site_id can exist with different subscope (unique constraint on site_id + subscope)
    - Sum column: Calculated automatically, not read from CSV
    - New site metadata: AC ARMOD Cable, Additional Cost, Remark, Partner, Request Status, etc.
    """
    logger.info(f"User {current_user.username} uploading OD BOQ CSV: {file.filename} for project {project_id}, consumed_year: {consumed_year}")

    # Check project access
    project = db.query(DUProject).filter(DUProject.pid_po == project_id).first()
    if not project:
        logger.warning(f"OD BOQ CSV upload attempted for non-existent project: {project_id}")
        raise HTTPException(status_code=404, detail="Project not found")

    if not check_du_project_access(current_user, project, db, "edit"):
        logger.warning(f"User {current_user.username} denied permission to upload OD BOQ for project {project_id}")
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You are not authorized to upload to this project."
        )

    if not file.filename.endswith('.csv'):
        logger.warning(f"Invalid file type uploaded for OD BOQ: {file.filename}")
        raise HTTPException(status_code=400, detail="File must be a CSV")

    try:
        content = await file.read()

        # Try different encodings
        csv_content = None
        for encoding in ['utf-8', 'utf-8-sig', 'cp1252', 'latin-1']:
            try:
                csv_content = content.decode(encoding)
                csv_content = csv_content.replace('\xa0', ' ')
                break
            except (UnicodeDecodeError, LookupError):
                continue

        if csv_content is None:
            raise HTTPException(status_code=400, detail="Could not decode CSV file")

        # Read CSV
        df = pd.read_csv(StringIO(csv_content), header=None)

        # Extract product headers (rows 1-6, starting from column 7)
        # Row 0: Description, Row 1: #Line, Row 2: Unit Price ($), Row 3: #Code,
        # Row 4: Category, Row 5: Total PO QTY, Row 6: Consumed (SKIP), Row 7: Remaining/Headers (SKIP)
        descriptions = df.iloc[0, 7:].tolist()
        line_numbers = df.iloc[1, 7:].tolist()
        unit_prices = df.iloc[2, 7:].tolist()
        codes = df.iloc[3, 7:].tolist()
        categories = df.iloc[4, 7:].tolist()
        total_pos = df.iloc[5, 7:].tolist()
        # Row 6 (index 6) is consumed in year - SKIP
        # Row 7 (index 7) is remaining in PO / column headers - SKIP

        # Helper function to safely extract and clean string values from lists
        def clean_value(lst, idx):
            """Extract value from list, return cleaned string or None."""
            if idx < len(lst) and not pd.isna(lst[idx]):
                val = str(lst[idx]).strip()
                return val if val else None
            return None

        # Create/update products (WITHOUT consumed_in_year and remaining_in_po)
        products_inserted = 0
        products_updated = 0
        product_id_map = {}  # Maps column index to product ID

        for idx, desc in enumerate(descriptions):
            # Skip columns without description OR without #Line number (validates product column)
            description = clean_value(descriptions, idx)
            line_number = clean_value(line_numbers, idx)

            if not description or not line_number:
                continue

            # Extract other product fields
            code = clean_value(codes, idx)
            category = clean_value(categories, idx)
            unit_price = parse_currency_value(unit_prices[idx]) if idx < len(unit_prices) else None
            total_po_qty = float(total_pos[idx]) if not pd.isna(total_pos[idx]) and str(total_pos[idx]).strip() else None

            # Check if product exists by code (upsert logic)
            existing_product = db.query(ODBOQProduct).filter(ODBOQProduct.code == code).first() if code else None

            if existing_product:
                # Update existing product (WITHOUT consumed_in_year and remaining_in_po)
                existing_product.description = description
                existing_product.line_number = line_number
                existing_product.category = category
                existing_product.unit_price = unit_price
                existing_product.total_po_qty = total_po_qty
                existing_product.consumed_year = consumed_year
                # consumed_in_year and remaining_in_po will be calculated later
                product_id_map[idx] = existing_product.id
                products_updated += 1
            else:
                # Create new product (WITHOUT consumed_in_year and remaining_in_po)
                new_product = ODBOQProduct(
                    description=description,
                    line_number=line_number,
                    code=code,
                    category=category,
                    unit_price=unit_price,
                    total_po_qty=total_po_qty,
                    consumed_year=consumed_year,
                    consumed_in_year=0,  # Will be calculated later
                    remaining_in_po=0  # Will be calculated later
                )
                db.add(new_product)
                db.flush()  # Get the ID without committing
                product_id_map[idx] = new_product.id
                products_inserted += 1

        # Determine number of product columns
        num_product_cols = len(product_id_map)

        # Calculate metadata column indices (after products and Sum column)
        # Structure: [Region, Distance, Scope, Subscope, Site ID, Model] + [Products...] + [Sum] + [Metadata...]
        metadata_start_col = 7 + num_product_cols + 1  # +1 for Sum column

        # Process site rows (row 8 onwards, which is index 7)
        sites_inserted = 0
        sites_updated = 0
        site_products_inserted = 0
        skipped = 0
        site_record_id_map = {}  # Maps (site_id, subscope) to database record ID

        # Helper function to safely extract cell value
        def safe_extract(row, col_idx):
            """Extract and clean cell value, return None if empty/invalid."""
            if col_idx < len(row) and not pd.isna(row.iloc[col_idx]):
                val = str(row.iloc[col_idx]).strip()
                return val if val else None
            return None

        for row_idx in range(8, len(df)):
            row = df.iloc[row_idx]

            # Extract site basic data (columns 0-5)
            region = safe_extract(row, 0)
            distance = safe_extract(row, 1)
            scope = safe_extract(row, 2)
            subscope = safe_extract(row, 3)
            site_id = safe_extract(row, 4)
            po_model = safe_extract(row, 5)

            # Skip if no site_id
            if not site_id or site_id == '':
                skipped += 1
                continue

            # Extract metadata columns (after products and Sum column) - 13 fields
            metadata_values = [safe_extract(row, metadata_start_col + i) for i in range(13)]
            (ac_armod_cable, additional_cost, remark, partner, request_status,
             requested_date, du_po_number, smp, year_scope, integration_status,
             integration_date, du_po_convention_name, po_year_issuance) = metadata_values

            # Check for existing site with same site_id and subscope
            existing_site = db.query(ODBOQSite).filter(
                ODBOQSite.site_id == site_id,
                ODBOQSite.subscope == subscope
            ).first()

            if existing_site:
                # Update existing site
                existing_site.region = region
                existing_site.distance = distance
                existing_site.scope = scope
                existing_site.po_model = po_model
                existing_site.project_id = project_id
                existing_site.ac_armod_cable = ac_armod_cable
                existing_site.additional_cost = additional_cost
                existing_site.remark = remark
                existing_site.partner = partner
                existing_site.request_status = request_status
                existing_site.requested_date = requested_date
                existing_site.du_po_number = du_po_number
                existing_site.smp = smp
                existing_site.year_scope = year_scope
                existing_site.integration_status = integration_status
                existing_site.integration_date = integration_date
                existing_site.du_po_convention_name = du_po_convention_name
                existing_site.po_year_issuance = po_year_issuance
                site_record_id_map[(site_id, subscope)] = existing_site.id
                sites_updated += 1
            else:
                # Create new site
                new_site = ODBOQSite(
                    site_id=site_id,
                    region=region,
                    distance=distance,
                    scope=scope,
                    subscope=subscope,
                    po_model=po_model,
                    project_id=project_id,
                    ac_armod_cable=ac_armod_cable,
                    additional_cost=additional_cost,
                    remark=remark,
                    partner=partner,
                    request_status=request_status,
                    requested_date=requested_date,
                    du_po_number=du_po_number,
                    smp=smp,
                    year_scope=year_scope,
                    integration_status=integration_status,
                    integration_date=integration_date,
                    du_po_convention_name=du_po_convention_name,
                    po_year_issuance=po_year_issuance
                )
                db.add(new_site)
                db.flush()  # Get the ID
                site_record_id_map[(site_id, subscope)] = new_site.id
                sites_inserted += 1

            # Get the site record ID for this site
            current_site_record_id = site_record_id_map[(site_id, subscope)]

            # Create site-product records (columns 7 to 7+num_product_cols-1)
            for col_idx, product_id in product_id_map.items():
                qty_value = row.iloc[7 + col_idx]
                qty = float(qty_value) if not pd.isna(qty_value) and str(qty_value).strip() != '' else None

                # Create or update site-product record
                existing_sp = db.query(ODBOQSiteProduct).filter(
                    ODBOQSiteProduct.site_record_id == current_site_record_id,
                    ODBOQSiteProduct.product_id == product_id
                ).first()

                if existing_sp:
                    existing_sp.qty_per_site = qty
                else:
                    new_sp = ODBOQSiteProduct(
                        site_record_id=current_site_record_id,
                        product_id=product_id,
                        qty_per_site=qty
                    )
                    db.add(new_sp)
                    site_products_inserted += 1

        # Flush to ensure all site-products are in the database
        db.flush()

        # Calculate consumed_in_year and remaining_in_po for each product (optimized batch query)
        product_ids = set(product_id_map.values())

        if product_ids:
            # Batch fetch all products at once
            products_dict = {
                p.id: p for p in db.query(ODBOQProduct).filter(
                    ODBOQProduct.id.in_(product_ids)
                ).all()
            }

            # Calculate all consumed amounts in a single query
            consumed_data = db.query(
                ODBOQSiteProduct.product_id,
                func.sum(ODBOQSiteProduct.qty_per_site).label('total')
            ).filter(
                ODBOQSiteProduct.product_id.in_(product_ids),
                ODBOQSiteProduct.qty_per_site.isnot(None)
            ).group_by(ODBOQSiteProduct.product_id).all()

            consumed_map = {cd[0]: cd[1] or 0.0 for cd in consumed_data}

            # Update products with calculated values
            for product_id, product in products_dict.items():
                total_consumed = consumed_map.get(product_id, 0.0)
                product.consumed_in_year = total_consumed

                # Calculate remaining_in_po as total_po_qty - consumed_in_year
                if product.total_po_qty is not None:
                    product.remaining_in_po = product.total_po_qty - total_consumed
                else:
                    product.remaining_in_po = -total_consumed  # Negative if no total_po_qty

        db.commit()

        logger.info(f"OD BOQ CSV upload completed: {sites_inserted} sites inserted, {sites_updated} updated, {products_inserted} products inserted, {products_updated} updated, {site_products_inserted} site-products inserted for project {project_id}")

        # Create audit log
        await create_audit_log(
            db=db,
            user_id=current_user.id,
            action="bulk_upload_od_boq",
            resource_type="od_boq_upload",
            resource_id="bulk",
            resource_name=file.filename,
            details=json.dumps({
                "project_id": project_id,
                "sites_inserted": sites_inserted,
                "sites_updated": sites_updated,
                "products_inserted": products_inserted,
                "products_updated": products_updated,
                "site_products_inserted": site_products_inserted
            }),
            ip_address=get_client_ip(request),
            user_agent=request.headers.get("User-Agent")
        )

        return UploadResponse(
            sites_inserted=sites_inserted,
            sites_updated=sites_updated,
            products_inserted=products_inserted,
            products_updated=products_updated,
            site_products_inserted=site_products_inserted,
            skipped=skipped,
            message=f"Successfully processed CSV: {sites_inserted} sites inserted, {products_inserted} products inserted"
        )

    except Exception as e:
        db.rollback()
        logger.error(f"Error processing OD BOQ CSV for project {project_id}: {str(e)}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error processing CSV: {str(e)}"
        )


# ===========================
# BOQ GENERATION ENDPOINTS
# ===========================

from Schemas.DU.OD_BOQ_Schema import (
    BOQGenerationRequest, BOQGenerationResult, BOQGenerationResponse,
    BOQExcelFromCSVRequest, BulkBOQEditedSiteData, BulkBOQExcelFromEditedRequest
)
from fastapi.responses import StreamingResponse
from io import BytesIO
import zipfile
from datetime import datetime
import os
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def generate_boq_csv_for_site(site: ODBOQSite, db: Session) -> str:
    """
    Generate BOQ CSV data for a single site.

    Logic:
    1. Query ODBOQSiteProduct joined with ODBOQProduct where qty_per_site > 0
    2. Apply category transformations (HW→1, SW→3, Service→1/2)
    3. Return CSV string with columns: Description, Category, BU, UOM, BOQ Qty, Line Number, Code
    """
    # Query site products with qty > 0
    site_products = db.query(
        ODBOQSiteProduct,
        ODBOQProduct
    ).join(
        ODBOQProduct, ODBOQSiteProduct.product_id == ODBOQProduct.id
    ).filter(
        ODBOQSiteProduct.site_record_id == site.id,
        ODBOQSiteProduct.qty_per_site > 0
    ).all()

    if not site_products:
        raise HTTPException(
            status_code=404,
            detail=f"No products found with qty_per_site > 0 for site '{site.site_id}'."
        )

    # Build result data - use raw quantities as stored
    result_data = []
    for sp, product in site_products:
        final_qty = sp.qty_per_site

        # Convert float to int if it's a whole number (e.g., 2.0 -> 2)
        if isinstance(final_qty, float) and final_qty == int(final_qty):
            final_qty = int(final_qty)

        up = product.unit_price
        total_usd = ''
        total_aed = ''
        if up is not None and final_qty is not None:
            total_usd_val = up * (final_qty if isinstance(final_qty, (int, float)) else 0)
            total_aed_val = total_usd_val * 3.6725
            total_usd = round(total_usd_val, 2)
            total_aed = round(total_aed_val, 2)

        result_data.append({
            'Description': product.description or ' ',
            'Category': product.category or ' ',
            'BU': product.bu or ' ',
            'UOM': product.code or ' ',
            'BOQ Qty': final_qty,
            'Unit Price': up if up is not None else ' ',
            'Total USD': total_usd if total_usd != '' else ' ',
            'Total AED': total_aed if total_aed != '' else ' ',
            'Line Number': product.line_number or ' ',
            'Code': product.code or ' '
        })

    # Get project details for header
    project = db.query(DUProject).filter(DUProject.pid_po == site.project_id).first()
    project_name = project.project_name if project else (site.project_id or 'N/A')
    project_po = project.po if project else (site.project_id or 'N/A')

    # Get current date
    current_date = datetime.now().strftime('%d-%b-%y')

    # Generate CSV string with metadata headers
    csv_lines = []

    # Add metadata header rows
    csv_lines.append(f'" "," "," "," ",DU BOQ," "," "')
    csv_lines.append(f'BPO Number:,{project_po}," "," "," ",Date:,{current_date}')
    csv_lines.append(f'Scope:,{site.scope or "N/A"}," "," "," ",Subscope:,{site.subscope or "N/A"}')
    csv_lines.append(f'Vendor:,Nokia," "," "," ",Site ID:,{site.site_id}')
    csv_lines.append(f'" "," "," "," "," "," "," "')  # Empty row separator

    # Add data table headers
    csv_headers = ['Description', 'Category', 'BU', 'UOM', 'BOQ Qty', 'Unit Price', 'Total USD', 'Total AED', 'Line Number', 'Code']
    csv_lines.append(','.join(csv_headers))

    # Add data rows
    for row in result_data:
        csv_row = []
        for header in csv_headers:
            value = row.get(header, ' ')
            value_str = str(value) if value is not None else ' '
            if value_str == '' or value_str == 'None':
                value_str = ' '
            # Escape commas and quotes in values
            if ',' in value_str or '"' in value_str:
                value_str = f'"{value_str.replace(chr(34), chr(34)+chr(34))}"'
            csv_row.append(value_str)
        csv_lines.append(','.join(csv_row))

    return '\n'.join(csv_lines)


@odBOQRoute.get("/sites/{site_record_id}/generate-boq")
async def generate_boq_for_site(
    site_record_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
) -> str:
    """
    Generate BOQ for a specific site.

    - site_record_id: The database ID (ODBOQSite.id)
    - Returns: CSV-formatted string with BOQ data
    """
    # Get the site
    site = db.query(ODBOQSite).filter(ODBOQSite.id == site_record_id).first()

    if not site:
        raise HTTPException(status_code=404, detail="Site not found")

    # Check project access if site has a project_id
    if site.project_id:
        project = db.query(DUProject).filter(DUProject.pid_po == site.project_id).first()
        if project and not check_du_project_access(current_user, project, db, "view"):
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You do not have access to this site's project."
            )

    # Generate BOQ CSV
    csv_content = generate_boq_csv_for_site(site, db)

    # Audit log
    await create_audit_log(
        db=db,
        user_id=current_user.id,
        action="generate_boq",
        resource_type="od_boq_site",
        resource_id=str(site_record_id),
        resource_name=site.site_id,
        details=json.dumps({"site_id": site.site_id, "subscope": site.subscope}),
        ip_address=get_client_ip(request),
        user_agent=request.headers.get("User-Agent")
    )

    return csv_content


@odBOQRoute.post("/sites/bulk-generate-boq", response_model=BOQGenerationResponse)
async def bulk_generate_boq(
    body: BOQGenerationRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Bulk generate BOQ for multiple sites.

    - body.site_record_ids: List of ODBOQSite.id values
    - Returns: BOQGenerationResponse with results for each site
    """
    results = []
    successful = 0
    failed = 0

    # Batch fetch all sites at once (fixes N+1 query)
    sites_map = {
        s.id: s for s in db.query(ODBOQSite).filter(
            ODBOQSite.id.in_(body.site_record_ids)
        ).all()
    }

    # Batch fetch all projects at once (fixes N+1 query)
    project_ids = [s.project_id for s in sites_map.values() if s.project_id]
    projects_map = {}
    if project_ids:
        projects_map = {
            p.pid_po: p for p in db.query(DUProject).filter(
                DUProject.pid_po.in_(project_ids)
            ).all()
        }

    for site_record_id in body.site_record_ids:
        site = sites_map.get(site_record_id)

        if not site:
            results.append(BOQGenerationResult(
                site_record_id=site_record_id,
                site_id="unknown",
                success=False,
                error="Site not found"
            ))
            failed += 1
            continue

        try:
            # Check project access using pre-fetched project
            if site.project_id:
                project = projects_map.get(site.project_id)
                if project and not check_du_project_access(current_user, project, db, "view"):
                    results.append(BOQGenerationResult(
                        site_record_id=site_record_id,
                        site_id=site.site_id,
                        subscope=site.subscope,
                        success=False,
                        error="Access denied"
                    ))
                    failed += 1
                    continue

            # Generate BOQ CSV
            csv_content = generate_boq_csv_for_site(site, db)

            results.append(BOQGenerationResult(
                site_record_id=site_record_id,
                site_id=site.site_id,
                subscope=site.subscope,
                success=True,
                csv_content=csv_content
            ))
            successful += 1

        except HTTPException as e:
            results.append(BOQGenerationResult(
                site_record_id=site_record_id,
                site_id=site.site_id,
                subscope=site.subscope,
                success=False,
                error=e.detail
            ))
            failed += 1
        except Exception as e:
            logger.error(f"Error generating BOQ for site {site.site_id}: {str(e)}", exc_info=True)
            results.append(BOQGenerationResult(
                site_record_id=site_record_id,
                site_id=site.site_id,
                subscope=site.subscope,
                success=False,
                error="Failed to generate BOQ"
            ))
            failed += 1

    # Audit log
    await create_audit_log(
        db=db,
        user_id=current_user.id,
        action="bulk_generate_boq",
        resource_type="od_boq_site",
        resource_id="bulk",
        details=json.dumps({
            "site_record_ids": body.site_record_ids,
            "successful": successful,
            "failed": failed
        }),
        ip_address=get_client_ip(request),
        user_agent=request.headers.get("User-Agent")
    )

    return BOQGenerationResponse(
        successful=successful,
        failed=failed,
        results=results
    )


def generate_boq_data_for_site(site: ODBOQSite, db: Session) -> dict:
    """
    Generate BOQ data for a single site in the format expected by the Excel template.
    Returns a dictionary with site metadata and BOQ items.
    """
    # Query site products with qty > 0
    site_products = db.query(
        ODBOQSiteProduct,
        ODBOQProduct
    ).join(
        ODBOQProduct, ODBOQSiteProduct.product_id == ODBOQProduct.id
    ).filter(
        ODBOQSiteProduct.site_record_id == site.id,
        ODBOQSiteProduct.qty_per_site > 0
    ).all()

    if not site_products:
        return {'error': f"No products found with qty_per_site > 0 for site '{site.site_id}'."}

    # Build result data - use raw quantities as stored
    result_data = []
    for sp, product in site_products:
        final_qty = sp.qty_per_site

        # Convert float to int if it's a whole number (e.g., 2.0 -> 2)
        if isinstance(final_qty, float) and final_qty == int(final_qty):
            final_qty = int(final_qty)

        # Calculate totals from unit_price if available
        up = product.unit_price
        total_usd = ''
        total_aed = ''
        if up is not None and final_qty is not None:
            total_usd_val = up * (final_qty if isinstance(final_qty, (int, float)) else 0)
            total_aed_val = total_usd_val * 3.6725  # USD to AED conversion
            total_usd = round(total_usd_val, 2)
            total_aed = round(total_aed_val, 2)

        result_data.append({
            'line': product.line_number,
            'bu': product.bu,
            'item_job': product.code,  # ERP Item Code
            'description': product.description,
            'budget_line': product.category,
            'qty': final_qty,
            'unit_price': up if up is not None else '',
            'total_usd': total_usd,
            'total_aed': total_aed,
            'site_id_list': site.site_id,
            'po_model': site.po_model or ''
        })

    # Get project details for header
    project = db.query(DUProject).filter(DUProject.pid_po == site.project_id).first()
    project_po = project.po if project else (site.project_id or 'N/A')

    return {
        'site_record_id': site.id,
        'site_id': site.site_id,
        'scope': site.scope,
        'subscope': site.subscope,
        'project_po': project_po,
        'sps_category': site.subscope or 'N/A',
        'po_model': site.po_model or '',
        'data': result_data
    }


def create_excel_from_boq_data(boq_entries: list, template_path: str, is_bulk: bool = False) -> BytesIO:
    """
    Create an Excel file from the template with BOQ data.

    Args:
        boq_entries: List of BOQ entry dictionaries from generate_boq_data_for_site
        template_path: Path to the Excel template file
        is_bulk: If True, combine all entries in one sheet

    Returns:
        BytesIO object containing the Excel file
    """
    # Load the template
    wb = load_workbook(template_path)
    ws = wb.active

    # Get max row and prepare template border
    max_row = ws.max_row

    # Get border style from template
    template_border = None
    if ws.max_row >= 9:
        for col in range(2, 13):
            sample_cell = ws.cell(row=9, column=col)
            if sample_cell.border and sample_cell.border.left and sample_cell.border.left.style:
                template_border = sample_cell.border.copy()
                break

    if template_border is None:
        thin_side = Side(style='thin', color='000000')
        template_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # Save footer rows
    footer_rows = []
    footer_start = None
    if max_row >= 10:
        for row_idx in range(max(10, max_row - 20), max_row + 1):
            has_fill = False
            for col_idx in range(1, 15):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb and cell.fill.start_color.rgb != '00000000':
                    has_fill = True
                    footer_start = row_idx
                    break
            if has_fill:
                break

        if footer_start:
            for row_idx in range(footer_start, max_row + 1):
                row_data = []
                for col_idx in range(1, 15):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    row_data.append({
                        'value': cell.value,
                        'font': cell.font.copy() if cell.font else None,
                        'fill': cell.fill.copy() if cell.fill else None,
                        'border': cell.border.copy() if cell.border else None,
                        'alignment': cell.alignment.copy() if cell.alignment else None
                    })
                footer_rows.append(row_data)
            ws.delete_rows(10, max_row - 9)

    if not is_bulk:
        # Single site mode
        if not boq_entries or len(boq_entries) == 0:
            return None

        entry = boq_entries[0]
        if not entry.get('data') or len(entry['data']) == 0:
            return None

        # Update header
        ws.cell(row=5, column=5).value = entry['project_po']
        ws.cell(row=5, column=9).value = datetime.now().strftime('%d-%b-%y')
        ws.cell(row=6, column=5).value = entry.get('scope', 'N/A')
        ws.cell(row=6, column=9).value = entry.get('sps_category', 'N/A')
        ws.cell(row=7, column=9).value = entry['site_id']

        start_row = 10
        sorted_data = sorted(entry['data'], key=lambda x: (x.get('line') is None, x.get('line') if x.get('line') is not None else 0))

        for idx, item in enumerate(sorted_data):
            row_num = start_row + idx

            # Column mapping
            ws.cell(row=row_num, column=1).value = ''
            ws.cell(row=row_num, column=1).border = template_border
            ws.cell(row=row_num, column=2).value = item.get('line')
            ws.cell(row=row_num, column=2).border = template_border
            ws.cell(row=row_num, column=3).value = item.get('bu')
            ws.cell(row=row_num, column=3).border = template_border
            ws.cell(row=row_num, column=4).value = item.get('item_job')
            ws.cell(row=row_num, column=4).border = template_border
            ws.cell(row=row_num, column=5).value = item.get('description')
            ws.cell(row=row_num, column=5).border = template_border
            ws.cell(row=row_num, column=6).value = ''
            ws.cell(row=row_num, column=6).border = template_border
            ws.cell(row=row_num, column=7).value = item.get('qty')
            ws.cell(row=row_num, column=7).border = template_border
            ws.cell(row=row_num, column=8).value = item.get('unit_price')
            ws.cell(row=row_num, column=8).border = template_border
            ws.cell(row=row_num, column=9).value = item.get('total_usd')
            ws.cell(row=row_num, column=9).border = template_border
            ws.cell(row=row_num, column=10).value = item.get('total_aed')
            ws.cell(row=row_num, column=10).border = template_border
            ws.cell(row=row_num, column=11).value = item.get('site_id_list')
            ws.cell(row=row_num, column=11).border = template_border
            ws.cell(row=row_num, column=12).value = item.get('po_model')
            ws.cell(row=row_num, column=12).border = template_border

        # Add totals row after data
        if len(sorted_data) > 0:
            last_data_row = start_row + len(sorted_data) - 1
            totals_row = last_data_row + 1

            thin_side = Side(style='thin', color='000000')
            totals_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
            bold_font = Font(bold=True)

            # Empty bordered cells for non-total columns
            for col in range(1, 13):
                cell = ws.cell(row=totals_row, column=col)
                cell.border = totals_border

            # "Total" label in description column
            ws.cell(row=totals_row, column=5).value = "Total"
            ws.cell(row=totals_row, column=5).font = bold_font

            # SUM formulas for Total USD (col 9) and Total AED (col 10)
            ws.cell(row=totals_row, column=9).value = f"=SUM(I{start_row}:I{last_data_row})"
            ws.cell(row=totals_row, column=9).font = bold_font
            ws.cell(row=totals_row, column=9).number_format = '#,##0.00'

            ws.cell(row=totals_row, column=10).value = f"=SUM(J{start_row}:J{last_data_row})"
            ws.cell(row=totals_row, column=10).font = bold_font
            ws.cell(row=totals_row, column=10).number_format = '#,##0.00'

        # Restore footer
        if footer_rows and len(sorted_data) > 0:
            footer_start_row = start_row + len(sorted_data) + 2  # +2 to account for totals row
            for footer_idx, footer_row_data in enumerate(footer_rows):
                row_num = footer_start_row + footer_idx
                for col_idx, cell_data in enumerate(footer_row_data):
                    cell = ws.cell(row=row_num, column=col_idx + 1)
                    cell.value = cell_data['value']
                    if cell_data['font']:
                        cell.font = cell_data['font']
                    if cell_data['fill']:
                        cell.fill = cell_data['fill']
                    if cell_data['border']:
                        cell.border = cell_data['border']
                    if cell_data['alignment']:
                        cell.alignment = cell_data['alignment']
    else:
        # Bulk mode
        if boq_entries and len(boq_entries) > 0:
            first_entry = boq_entries[0]
            ws.cell(row=5, column=5).value = first_entry['project_po']
            ws.cell(row=5, column=9).value = datetime.now().strftime('%d-%b-%y')
            ws.cell(row=6, column=5).value = "Multiple Scopes (Bulk)"
            ws.cell(row=6, column=9).value = "Multiple Sites"
            ws.cell(row=7, column=9).value = f"{len(boq_entries)} sites"

        all_data = []
        for entry in boq_entries:
            for item in entry.get('data', []):
                item_with_site = item.copy()
                item_with_site['site_id_list'] = entry['site_id']
                item_with_site['po_model'] = entry.get('po_model', '')
                all_data.append(item_with_site)

        # Sort by line number numerically (convert string to int for proper sorting)
        def get_line_number_key(x):
            line = x.get('line')
            if line is None:
                return (True, 0)  # None values go last
            try:
                return (False, int(line))
            except (ValueError, TypeError):
                return (False, 0)  # Non-numeric values sorted as 0
        sorted_data = sorted(all_data, key=get_line_number_key)
        start_row = 10

        for idx, item in enumerate(sorted_data):
            row_num = start_row + idx

            ws.cell(row=row_num, column=1).value = ''
            ws.cell(row=row_num, column=1).border = template_border
            ws.cell(row=row_num, column=2).value = item.get('line')
            ws.cell(row=row_num, column=2).border = template_border
            ws.cell(row=row_num, column=3).value = item.get('bu')
            ws.cell(row=row_num, column=3).border = template_border
            ws.cell(row=row_num, column=4).value = item.get('item_job')
            ws.cell(row=row_num, column=4).border = template_border
            ws.cell(row=row_num, column=5).value = item.get('description')
            ws.cell(row=row_num, column=5).border = template_border
            ws.cell(row=row_num, column=6).value = ''
            ws.cell(row=row_num, column=6).border = template_border
            ws.cell(row=row_num, column=7).value = item.get('qty')
            ws.cell(row=row_num, column=7).border = template_border
            ws.cell(row=row_num, column=8).value = item.get('unit_price')
            ws.cell(row=row_num, column=8).border = template_border
            ws.cell(row=row_num, column=9).value = item.get('total_usd')
            ws.cell(row=row_num, column=9).border = template_border
            ws.cell(row=row_num, column=10).value = item.get('total_aed')
            ws.cell(row=row_num, column=10).border = template_border
            ws.cell(row=row_num, column=11).value = item.get('site_id_list')
            ws.cell(row=row_num, column=11).border = template_border
            ws.cell(row=row_num, column=12).value = item.get('po_model')
            ws.cell(row=row_num, column=12).border = template_border

        # Add totals row after data
        if len(sorted_data) > 0:
            last_data_row = start_row + len(sorted_data) - 1
            totals_row = last_data_row + 1

            thin_side = Side(style='thin', color='000000')
            totals_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
            bold_font = Font(bold=True)

            # Empty bordered cells for non-total columns
            for col in range(1, 13):
                cell = ws.cell(row=totals_row, column=col)
                cell.border = totals_border

            # "Total" label in description column
            ws.cell(row=totals_row, column=5).value = "Total"
            ws.cell(row=totals_row, column=5).font = bold_font

            # SUM formulas for Total USD (col 9) and Total AED (col 10)
            ws.cell(row=totals_row, column=9).value = f"=SUM(I{start_row}:I{last_data_row})"
            ws.cell(row=totals_row, column=9).font = bold_font
            ws.cell(row=totals_row, column=9).number_format = '#,##0.00'

            ws.cell(row=totals_row, column=10).value = f"=SUM(J{start_row}:J{last_data_row})"
            ws.cell(row=totals_row, column=10).font = bold_font
            ws.cell(row=totals_row, column=10).number_format = '#,##0.00'

        # Restore footer
        if footer_rows and len(sorted_data) > 0:
            footer_start_row = start_row + len(sorted_data) + 2  # +2 to account for totals row
            for footer_idx, footer_row_data in enumerate(footer_rows):
                row_num = footer_start_row + footer_idx
                for col_idx, cell_data in enumerate(footer_row_data):
                    cell = ws.cell(row=row_num, column=col_idx + 1)
                    cell.value = cell_data['value']
                    if cell_data['font']:
                        cell.font = cell_data['font']
                    if cell_data['fill']:
                        cell.fill = cell_data['fill']
                    if cell_data['border']:
                        cell.border = cell_data['border']
                    if cell_data['alignment']:
                        cell.alignment = cell_data['alignment']

    # --- Column widths and formatting ---
    # Set column widths so data is visible without manual resizing
    column_widths = {
        'A': 4,    # Empty/index
        'B': 12,   # BPO Line No
        'C': 14,   # BU
        'D': 18,   # ERP Item Code
        'E': 45,   # Description (widest)
        'F': 14,   # Budget Line
        'G': 10,   # QTY
        'H': 14,   # Unit Price
        'I': 16,   # Total USD
        'J': 16,   # Total AED
        'K': 16,   # Site ID
        'L': 30,   # Model Name towards dU
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Apply text wrapping to all data cells so content is visible
    wrap_alignment = Alignment(wrap_text=True, vertical='center')
    data_start = 10
    data_end = ws.max_row
    for row_idx in range(data_start, data_end + 1):
        for col_idx in range(1, 13):
            cell = ws.cell(row=row_idx, column=col_idx)
            # Preserve existing alignment properties but add wrap_text
            existing = cell.alignment
            cell.alignment = Alignment(
                horizontal=existing.horizontal if existing.horizontal else None,
                vertical='center',
                wrap_text=True
            )

    # Number formatting for currency columns in data rows
    for row_idx in range(data_start, data_end + 1):
        for col_idx in [8, 9, 10]:  # Unit Price, Total USD, Total AED
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@odBOQRoute.get("/sites/{site_record_id}/download-boq-excel")
async def download_boq_excel(
    site_record_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Download BOQ as Excel file for a specific site.

    - site_record_id: The database ID (ODBOQSite.id)
    - Returns: Excel file as StreamingResponse
    """
    # Get the site
    site = db.query(ODBOQSite).filter(ODBOQSite.id == site_record_id).first()

    if not site:
        raise HTTPException(status_code=404, detail="Site not found")

    # Check project access
    if site.project_id:
        project = db.query(DUProject).filter(DUProject.pid_po == site.project_id).first()
        if project and not check_du_project_access(current_user, project, db, "view"):
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You do not have access to this site's project."
            )

    # Generate BOQ data
    boq_data = generate_boq_data_for_site(site, db)

    if 'error' in boq_data:
        raise HTTPException(status_code=404, detail=boq_data['error'])

    # Get template path
    template_path = os.path.join(os.path.dirname(__file__), '..', '..', 'templates', 'BOQ Formate.xlsx')

    if not os.path.exists(template_path):
        raise HTTPException(
            status_code=500,
            detail="Excel template file not found. Please contact administrator."
        )

    # Create Excel from template
    excel_file = create_excel_from_boq_data([boq_data], template_path, is_bulk=False)

    if excel_file is None:
        raise HTTPException(status_code=500, detail="Failed to generate Excel file")

    # Create filename
    date_str = datetime.now().strftime('%Y-%m-%d')
    filename = f"BOQ_{site.site_id}_{date_str}.xlsx"

    # Audit log
    await create_audit_log(
        db=db,
        user_id=current_user.id,
        action="download_boq_excel",
        resource_type="od_boq_site",
        resource_id=str(site_record_id),
        resource_name=site.site_id,
        details=json.dumps({"site_id": site.site_id, "subscope": site.subscope, "filename": filename}),
        ip_address=get_client_ip(request),
        user_agent=request.headers.get("User-Agent")
    )

    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@odBOQRoute.post("/download-boq-excel-from-csv")
async def download_boq_excel_from_csv(
    body: BOQExcelFromCSVRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Download BOQ as Excel file from edited CSV data.

    - body.site_record_id: The database ID
    - body.site_id: The site ID string (for filename)
    - body.csv_data: 2D array of CSV data
    - Returns: Excel file as StreamingResponse
    """
    # Create DataFrame from 2D array
    df = pd.DataFrame(body.csv_data)

    # Write to Excel
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='BOQ', index=False, header=False)
    output.seek(0)

    # Create filename
    date_str = datetime.now().strftime('%Y-%m-%d')
    filename = f"BOQ_{body.site_id}_{date_str}.xlsx"

    # Audit log
    await create_audit_log(
        db=db,
        user_id=current_user.id,
        action="download_boq_excel_from_csv",
        resource_type="od_boq_site",
        resource_id=str(body.site_record_id),
        resource_name=body.site_id,
        details=json.dumps({"site_id": body.site_id, "filename": filename, "rows": len(body.csv_data)}),
        ip_address=get_client_ip(request),
        user_agent=request.headers.get("User-Agent")
    )

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@odBOQRoute.post("/sites/bulk-download-boq-zip")
async def bulk_download_boq_excel(
    body: BOQGenerationRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Bulk download BOQ as a single Excel file with all sites combined, ordered by BPO Line No.

    - body.site_record_ids: List of ODBOQSite.id values
    - Returns: Single Excel file as StreamingResponse
    """
    # Get template path
    template_path = os.path.join(os.path.dirname(__file__), '..', '..', 'templates', 'BOQ Formate.xlsx')

    if not os.path.exists(template_path):
        raise HTTPException(
            status_code=500,
            detail="Excel template file not found. Please contact administrator."
        )

    # Collect all BOQ data from all sites
    all_boq_entries = []

    for site_record_id in body.site_record_ids:
        try:
            # Get the site
            site = db.query(ODBOQSite).filter(ODBOQSite.id == site_record_id).first()

            if not site:
                continue

            # Check project access
            if site.project_id:
                project = db.query(DUProject).filter(DUProject.pid_po == site.project_id).first()
                if project and not check_du_project_access(current_user, project, db, "view"):
                    continue

            # Generate BOQ data
            boq_data = generate_boq_data_for_site(site, db)

            if 'error' in boq_data:
                logger.warning(f"No BOQ data for site {site_record_id}: {boq_data['error']}")
                continue

            all_boq_entries.append(boq_data)

        except Exception as e:
            logger.warning(f"Failed to generate BOQ for site {site_record_id}: {str(e)}")
            continue

    if not all_boq_entries:
        raise HTTPException(
            status_code=404,
            detail="No BOQ data found for any of the selected sites."
        )

    # Create single Excel from template with all entries combined (is_bulk=True)
    excel_file = create_excel_from_boq_data(all_boq_entries, template_path, is_bulk=True)

    if excel_file is None:
        raise HTTPException(status_code=500, detail="Failed to generate Excel file")

    # Create filename
    date_str = datetime.now().strftime('%Y-%m-%d')
    filename = f"BOQ_Bulk_{len(all_boq_entries)}_sites_{date_str}.xlsx"

    # Audit log
    await create_audit_log(
        db=db,
        user_id=current_user.id,
        action="bulk_download_boq_excel",
        resource_type="od_boq_site",
        resource_id="bulk",
        details=json.dumps({
            "site_record_ids": body.site_record_ids,
            "sites_included": len(all_boq_entries),
            "filename": filename
        }),
        ip_address=get_client_ip(request),
        user_agent=request.headers.get("User-Agent")
    )

    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@odBOQRoute.post("/sites/bulk-download-boq-excel-from-edited")
async def bulk_download_boq_excel_from_edited(
    body: BulkBOQExcelFromEditedRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Bulk download BOQ as a single Excel file from edited CSV data.
    This endpoint accepts the modified data from the frontend modal.

    - body.sites_data: List of edited site data with csv_data arrays
    - Returns: Single Excel file as StreamingResponse
    """
    try:
        # Get template path
        template_path = os.path.join(os.path.dirname(__file__), '..', '..', 'templates', 'BOQ Formate.xlsx')

        if not os.path.exists(template_path):
            raise HTTPException(
                status_code=500,
                detail="Excel template file not found. Please contact administrator."
            )

        if not body.sites_data:
            raise HTTPException(
                status_code=400,
                detail="No site data provided."
            )

        # Convert edited CSV data to BOQ entries format expected by create_excel_from_boq_data
        all_boq_entries = []

        for site_data in body.sites_data:
            csv_data = site_data.csv_data
            # Skip header rows (first 6 rows are metadata/headers in the CSV modal)
            # Find where the data rows start (after "Description" header row)
            data_start_idx = 6  # Default: data starts at row 6 (after 5 metadata + 1 header)

            for i, row in enumerate(csv_data):
                if row and len(row) > 0:
                    # Check for Description header (strip whitespace for comparison)
                    first_cell = str(row[0]).strip()
                    if first_cell == 'Description':
                        data_start_idx = i + 1
                        break

            # Extract data rows
            boq_items = []
            for row in csv_data[data_start_idx:]:
                if row and len(row) >= 1:
                    # CSV format: Description, Category, BU, UOM, BOQ Qty, Unit Price, Total USD, Total AED, Line Number, Code
                    description = str(row[0]).strip() if len(row) > 0 else ''
                    category = str(row[1]).strip() if len(row) > 1 else ''
                    bu = str(row[2]).strip() if len(row) > 2 else ''
                    uom = str(row[3]).strip() if len(row) > 3 else ''
                    qty_str = str(row[4]).strip() if len(row) > 4 else ''
                    unit_price_str = str(row[5]).strip() if len(row) > 5 else ''
                    total_usd_str = str(row[6]).strip() if len(row) > 6 else ''
                    total_aed_str = str(row[7]).strip() if len(row) > 7 else ''
                    line_number = str(row[8]).strip() if len(row) > 8 else ''
                    code = str(row[9]).strip() if len(row) > 9 else ''

                    # Skip empty rows
                    if not description and not code:
                        continue

                    # Parse quantity
                    try:
                        qty = float(qty_str) if qty_str else 0
                    except (ValueError, TypeError):
                        qty = 0

                    # Parse unit price
                    try:
                        unit_price = float(unit_price_str) if unit_price_str else ''
                    except (ValueError, TypeError):
                        unit_price = ''

                    # Parse totals
                    try:
                        total_usd = float(total_usd_str) if total_usd_str else ''
                    except (ValueError, TypeError):
                        total_usd = ''

                    try:
                        total_aed = float(total_aed_str) if total_aed_str else ''
                    except (ValueError, TypeError):
                        total_aed = ''

                    boq_items.append({
                        'line': line_number,
                        'bu': bu,
                        'item_job': code,
                        'description': description,
                        'qty': qty,
                        'unit_price': unit_price,
                        'total_usd': total_usd,
                        'total_aed': total_aed
                    })

            if boq_items:
                all_boq_entries.append({
                    'site_id': site_data.site_id,
                    'subscope': site_data.subscope,
                    'po_model': '',  # Not available in edited data
                    'project_po': 'N/A',  # Required by create_excel_from_boq_data
                    'scope': site_data.subscope or 'N/A',
                    'sps_category': site_data.subscope or 'N/A',
                    'data': boq_items
                })

        if not all_boq_entries:
            raise HTTPException(
                status_code=400,
                detail="No valid BOQ data found in the edited data."
            )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error processing bulk edited data: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Error processing data. Please check your input and try again.")

    try:
        # Create single Excel from template with all entries combined (is_bulk=True)
        excel_file = create_excel_from_boq_data(all_boq_entries, template_path, is_bulk=True)

        if excel_file is None:
            raise HTTPException(status_code=500, detail="Failed to generate Excel file")

        # Create filename
        date_str = datetime.now().strftime('%Y-%m-%d')
        filename = f"BOQ_Bulk_{len(all_boq_entries)}_sites_{date_str}.xlsx"

        # Audit log
        await create_audit_log(
            db=db,
            user_id=current_user.id,
            action="bulk_download_boq_excel_from_edited",
            resource_type="od_boq_site",
            resource_id="bulk_edited",
            details=json.dumps({
                "sites_count": len(body.sites_data),
                "sites_included": len(all_boq_entries),
                "filename": filename
            }),
            ip_address=get_client_ip(request),
            user_agent=request.headers.get("User-Agent")
        )

        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generating Excel from edited data: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Error generating Excel file. Please try again.")
