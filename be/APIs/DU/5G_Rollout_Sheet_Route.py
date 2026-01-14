# routes/5G_Rollout_Sheet_Route.py
"""
5G Rollout Sheet API Routes

This module provides CRUD operations for managing 5G Rollout Sheet data.
It includes pagination, search, admin control, and CSV upload functionality.
"""

import json
import csv
import os
from io import StringIO, BytesIO
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query, status, Request, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_, func
from typing import Optional, List
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

from APIs.Core import get_db, get_current_user
from Models.Admin.User import User, UserProjectAccess
from Models.Admin.AuditLog import AuditLog
from Models.DU.OD_BOQ_Item import ODBOQItem, LEVEL1_CATEGORIES, QUANTITY_FIELDS
from Models.DU.CustomerPO import CustomerPO

# Import model using importlib since filename starts with number
import importlib
rollout_model = importlib.import_module("Models.DU.5G_Rollout_Sheet")
_5G_Rollout_Sheet = rollout_model._5G_Rollout_Sheet

# Import DU Project model
du_project_model = importlib.import_module("Models.DU.DU_Project")
DUProject = du_project_model.DUProject

# Import schemas using importlib since filename starts with number
rollout_schema = importlib.import_module("Schemas.DU.5G_Rollout_Sheet_Schema")
Create5GRolloutSheet = rollout_schema.Create5GRolloutSheet
Update5GRolloutSheet = rollout_schema.Update5GRolloutSheet
RolloutSheetOut = rollout_schema.RolloutSheetOut
RolloutSheetPagination = rollout_schema.RolloutSheetPagination
RolloutSheetStatsResponse = rollout_schema.RolloutSheetStatsResponse
UploadResponse = rollout_schema.UploadResponse

rolloutSheetRoute = APIRouter(tags=["5G Rollout Sheet"])


# ===========================
# HELPER FUNCTIONS
# ===========================

async def create_audit_log(
        db: Session,
        user_id: int,
        action: str,
        resource_type: str,
        resource_id: Optional[str] = None,
        resource_name: Optional[str] = None,
        details: Optional[str] = None,
        ip_address: Optional[str] = None,
        user_agent: Optional[str] = None
):
    """Create an audit log entry."""
    audit_log = AuditLog(
        user_id=user_id,
        action=action,
        resource_type=resource_type,
        resource_id=resource_id,
        resource_name=resource_name,
        details=details,
        ip_address=ip_address,
        user_agent=user_agent
    )
    db.add(audit_log)
    db.commit()
    return audit_log


def get_client_ip(request: Request) -> str:
    """Extract client IP from request."""
    forwarded = request.headers.get("X-Forwarded-For")
    if forwarded:
        return forwarded.split(",")[0].strip()
    return request.client.host if request.client else "unknown"


def check_admin_access(current_user: User, required_permission: str = "view"):
    """
    Check if user has admin access with required permission level.

    Args:
        current_user: The current user
        required_permission: Required permission level ("view", "edit", "all")

    Returns:
        bool: True if user has access, False otherwise
    """
    # Senior admin has all permissions
    if current_user.role.name == "senior_admin":
        return True

    # Admin has edit and view permissions
    if current_user.role.name == "admin":
        if required_permission in ["view", "edit"]:
            return True

    # Regular users only have view permission
    if required_permission == "view":
        return True

    return False


def filter_rollout_by_user_access(current_user: User, query, db: Session):
    """
    Filter rollout sheet query based on user's project access.
    Senior admins can see all records.
    """
    if current_user.role.name == "senior_admin":
        return query

    # Get accessible project IDs for non-senior admins
    user_accesses = db.query(UserProjectAccess).filter(
        UserProjectAccess.user_id == current_user.id
    ).all()

    if not user_accesses:
        # User has no project access, return empty query
        return query.filter(_5G_Rollout_Sheet.id == -1)

    accessible_project_ids = [access.project_id for access in user_accesses]

    # Filter by accessible projects
    return query.filter(_5G_Rollout_Sheet.project_id.in_(accessible_project_ids))


# ===========================
# CRUD OPERATIONS
# ===========================

@rolloutSheetRoute.post("/rollout-sheet", response_model=RolloutSheetOut)
async def create_rollout_sheet(
        data: Create5GRolloutSheet,
        request: Request,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    """
    Create a new 5G Rollout Sheet entry.
    Requires 'edit' or 'all' permission.
    """
    # Check permission
    if not check_admin_access(current_user, "edit"):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You are not authorized to create rollout sheet entries. Contact the Senior Admin."
        )

    try:
        new_entry = _5G_Rollout_Sheet(**data.dict())
        db.add(new_entry)
        db.commit()
        db.refresh(new_entry)

        # Create audit log
        await create_audit_log(
            db=db,
            user_id=current_user.id,
            action="create_rollout_sheet",
            resource_type="rollout_sheet",
            resource_id=str(new_entry.id),
            resource_name=new_entry.site_id,
            details=json.dumps({
                "site_id": data.site_id,
                "partner": data.partner,
                "request_status": data.request_status
            }),
            ip_address=get_client_ip(request),
            user_agent=request.headers.get("User-Agent")
        )

        return new_entry
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error creating rollout sheet entry: {str(e)}"
        )


@rolloutSheetRoute.get("/rollout-sheet/stats", response_model=RolloutSheetStatsResponse)
def get_rollout_sheet_stats(
        project_id: Optional[str] = None,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    """
    Get rollout sheet statistics (total items, unique sites, unique partners).
    Optionally filter by project_id.
    """
    try:
        query = db.query(_5G_Rollout_Sheet)

        # Filter by user access
        query = filter_rollout_by_user_access(current_user, query, db)

        # Filter by project if specified
        if project_id:
            query = query.filter(_5G_Rollout_Sheet.project_id == project_id)

        total_items = query.count()

        # Count unique sites
        unique_sites = query.with_entities(
            func.count(func.distinct(_5G_Rollout_Sheet.site_id))
        ).scalar() or 0

        # Count unique partners
        unique_partners = query.with_entities(
            func.count(func.distinct(_5G_Rollout_Sheet.partner))
        ).scalar() or 0

        return RolloutSheetStatsResponse(
            total_items=total_items,
            unique_sites=unique_sites,
            unique_partners=unique_partners
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error retrieving rollout sheet stats: {str(e)}"
        )


@rolloutSheetRoute.get("/rollout-sheet", response_model=RolloutSheetPagination)
def get_rollout_sheets(
        skip: int = Query(0, ge=0),
        limit: int = Query(50, ge=1, le=500),
        search: Optional[str] = Query(None),
        project_id: Optional[str] = Query(None),
        partner: Optional[str] = Query(None),
        request_status: Optional[str] = Query(None),
        integration_status: Optional[str] = Query(None),
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    """
    Get paginated list of 5G Rollout Sheet entries with optional filters.
    Users can only see entries from projects they have access to.
    """
    try:
        query = db.query(_5G_Rollout_Sheet)

        # Filter by user access
        query = filter_rollout_by_user_access(current_user, query, db)

        # Apply filters
        if project_id:
            query = query.filter(_5G_Rollout_Sheet.project_id == project_id)

        if partner:
            query = query.filter(_5G_Rollout_Sheet.partner == partner)

        if request_status:
            query = query.filter(_5G_Rollout_Sheet.request_status == request_status)

        if integration_status:
            query = query.filter(_5G_Rollout_Sheet.integration_status == integration_status)

        # Apply search across multiple fields
        if search:
            search_pattern = f"%{search}%"
            query = query.filter(
                or_(
                    _5G_Rollout_Sheet.site_id.ilike(search_pattern),
                    _5G_Rollout_Sheet.partner.ilike(search_pattern),
                    _5G_Rollout_Sheet.partner_requester_name.ilike(search_pattern),
                    _5G_Rollout_Sheet.du_po_number.ilike(search_pattern),
                    _5G_Rollout_Sheet.smp_number.ilike(search_pattern),
                    _5G_Rollout_Sheet.wo_number.ilike(search_pattern),
                    _5G_Rollout_Sheet.nokia_rollout_requester.ilike(search_pattern)
                )
            )

        total_count = query.count()
        records = query.order_by(_5G_Rollout_Sheet.id.desc()).offset(skip).limit(limit).all()

        return RolloutSheetPagination(records=records, total=total_count)
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error retrieving rollout sheets: {str(e)}"
        )


@rolloutSheetRoute.get("/rollout-sheet/filters/options")
def get_filter_options(
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    """
    Get unique values for filter dropdowns (partners, request statuses, integration statuses).
    Must be defined before /{entry_id} to avoid route conflict.
    """
    try:
        query = db.query(_5G_Rollout_Sheet)
        query = filter_rollout_by_user_access(current_user, query, db)

        # Get unique partners
        partners = query.with_entities(
            _5G_Rollout_Sheet.partner
        ).distinct().all()
        partners = [p[0] for p in partners if p[0]]

        # Get unique request statuses
        request_statuses = query.with_entities(
            _5G_Rollout_Sheet.request_status
        ).distinct().all()
        request_statuses = [s[0] for s in request_statuses if s[0]]

        # Get unique integration statuses
        integration_statuses = query.with_entities(
            _5G_Rollout_Sheet.integration_status
        ).distinct().all()
        integration_statuses = [s[0] for s in integration_statuses if s[0]]

        # Get unique year target scopes
        year_scopes = query.with_entities(
            _5G_Rollout_Sheet.year_target_scope
        ).distinct().all()
        year_scopes = [y[0] for y in year_scopes if y[0]]

        return {
            "partners": sorted(partners),
            "request_statuses": sorted(request_statuses),
            "integration_statuses": sorted(integration_statuses),
            "year_target_scopes": sorted(year_scopes)
        }
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error retrieving filter options: {str(e)}"
        )


@rolloutSheetRoute.get("/rollout-sheet/{entry_id}", response_model=RolloutSheetOut)
def get_rollout_sheet_by_id(
        entry_id: int,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    """
    Get a specific 5G Rollout Sheet entry by ID.
    """
    entry = db.query(_5G_Rollout_Sheet).filter(_5G_Rollout_Sheet.id == entry_id).first()

    if not entry:
        raise HTTPException(status_code=404, detail="Rollout sheet entry not found")

    # Check if user has access to this entry's project
    if current_user.role.name != "senior_admin" and entry.project_id:
        user_access = db.query(UserProjectAccess).filter(
            UserProjectAccess.user_id == current_user.id,
            UserProjectAccess.project_id == entry.project_id
        ).first()

        if not user_access:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You do not have access to this rollout sheet entry."
            )

    return entry


@rolloutSheetRoute.put("/rollout-sheet/{entry_id}", response_model=RolloutSheetOut)
async def update_rollout_sheet(
        entry_id: int,
        data: Update5GRolloutSheet,
        request: Request,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    """
    Update an existing 5G Rollout Sheet entry.
    Requires 'edit' or 'all' permission.
    """
    # Check permission
    if not check_admin_access(current_user, "edit"):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You are not authorized to update rollout sheet entries. Contact the Senior Admin."
        )

    entry = db.query(_5G_Rollout_Sheet).filter(_5G_Rollout_Sheet.id == entry_id).first()

    if not entry:
        raise HTTPException(status_code=404, detail="Rollout sheet entry not found")

    try:
        # Store old data for audit log
        old_data = {
            "site_id": entry.site_id,
            "partner": entry.partner,
            "request_status": entry.request_status
        }

        # Update only provided fields
        update_data = data.dict(exclude_unset=True)
        for field, value in update_data.items():
            if value is not None:
                setattr(entry, field, value)

        db.commit()
        db.refresh(entry)

        # Create audit log
        await create_audit_log(
            db=db,
            user_id=current_user.id,
            action="update_rollout_sheet",
            resource_type="rollout_sheet",
            resource_id=str(entry.id),
            resource_name=entry.site_id,
            details=json.dumps({
                "old_data": old_data,
                "new_data": update_data
            }),
            ip_address=get_client_ip(request),
            user_agent=request.headers.get("User-Agent")
        )

        return entry
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error updating rollout sheet entry: {str(e)}"
        )


@rolloutSheetRoute.delete("/rollout-sheet/{entry_id}")
async def delete_rollout_sheet(
        entry_id: int,
        request: Request,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    """
    Delete a 5G Rollout Sheet entry.
    Requires 'all' permission (senior_admin only).
    """
    # Check permission - only senior_admin can delete
    if not check_admin_access(current_user, "all"):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You are not authorized to delete rollout sheet entries. Contact the Senior Admin."
        )

    entry = db.query(_5G_Rollout_Sheet).filter(_5G_Rollout_Sheet.id == entry_id).first()

    if not entry:
        raise HTTPException(status_code=404, detail="Rollout sheet entry not found")

    try:
        # Store data for audit log before deletion
        entry_data = {
            "site_id": entry.site_id,
            "partner": entry.partner,
            "project_id": entry.project_id,
            "request_status": entry.request_status
        }

        db.delete(entry)
        db.commit()

        # Create audit log
        await create_audit_log(
            db=db,
            user_id=current_user.id,
            action="delete_rollout_sheet",
            resource_type="rollout_sheet",
            resource_id=str(entry_id),
            resource_name=entry_data["site_id"],
            details=json.dumps(entry_data),
            ip_address=get_client_ip(request),
            user_agent=request.headers.get("User-Agent")
        )

        return {"message": f"Rollout sheet entry {entry_id} deleted successfully"}
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error deleting rollout sheet entry: {str(e)}"
        )


# ===========================
# BULK OPERATIONS
# ===========================

@rolloutSheetRoute.post("/rollout-sheet/upload-csv", response_model=UploadResponse)
async def upload_rollout_sheet_csv(
        file: UploadFile = File(...),
        project_id: str = Form(None),
        request: Request = None,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    """
    Upload a CSV file to create/update 5G Rollout Sheet entries.
    Requires 'edit' or 'all' permission.
    """
    # Check permission
    if not check_admin_access(current_user, "edit"):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You are not authorized to upload rollout sheet data. Contact the Senior Admin."
        )

    if not file.filename.endswith('.csv'):
        raise HTTPException(status_code=400, detail="File must be a CSV")

    try:
        content = await file.read()

        # Try multiple encodings to handle different CSV formats
        csv_content = None
        encodings = ['utf-8', 'cp1252', 'latin-1', 'iso-8859-1']

        for encoding in encodings:
            try:
                csv_content = StringIO(content.decode(encoding))
                break
            except UnicodeDecodeError:
                continue

        if csv_content is None:
            raise HTTPException(
                status_code=400,
                detail="Unable to decode CSV file. Please ensure it's saved in UTF-8, Windows-1252, or Latin-1 encoding."
            )

        csv_reader = csv.DictReader(csv_content)

        # Define header mapping for CSV columns to model fields
        header_mapping = {
            'SiteID': 'site_id',
            'Scope': 'scope',
            'Year Target Scope': 'year_target_scope',
            'Partner': 'partner',
            'Partner Requester Name': 'partner_requester_name',
            'Date of Partner Request': 'date_of_partner_request',
            'Survey Partner': 'survey_partner',
            'Implementation Partner': 'implementation_partner',
            'Ant Swap': 'ant_swap',
            'Additional Cost': 'additional_cost',
            'WR Transportation': 'wr_transportation',
            'Crane': 'crane',
            'AC Armod Cable New SRAN': 'ac_armod_cable_new_sran',
            'Military Factor': 'military_factor',
            'CICPA Factor': 'cicpa_factor',
            'Nokia Rollout Requester': 'nokia_rollout_requester',
            'Services Validation by Rollout': 'services_validation_by_rollout',
            'Date of Validation by Rollout': 'date_of_validation_by_rollout',
            'Request Status': 'request_status',
            'Du PO#': 'du_po_number',
            'Integration Status': 'integration_status',
            'Integration Date': 'integration_date',
            'DU PO Convention Name': 'du_po_convention_name',
            'PO Year Issuance': 'po_year_issuance',
            'SMP#': 'smp_number',
            'WO#': 'wo_number',
            'SPS Category': 'sps_category',
            'Submission Date': 'submission_date',
            'PO Status': 'po_status',
            'PAC Received': 'pac_received',
            'Date of PAC': 'date_of_pac',
            'Hardware Remark': 'hardware_remark',
            'Project ID': 'project_id'
        }

        inserted_count = 0
        updated_count = 0

        for row in csv_reader:
            # Map CSV columns to model fields
            entry_data = {}
            for csv_header, model_field in header_mapping.items():
                value = row.get(csv_header, '').strip() if row.get(csv_header) else None
                if value:
                    entry_data[model_field] = value

            # Use form project_id if not in CSV
            if not entry_data.get('project_id') and project_id:
                entry_data['project_id'] = project_id

            # Skip rows without site_id
            if not entry_data.get('site_id'):
                continue

            # Check if entry exists (by site_id and project_id)
            existing_entry = db.query(_5G_Rollout_Sheet).filter(
                _5G_Rollout_Sheet.site_id == entry_data['site_id'],
                _5G_Rollout_Sheet.project_id == entry_data.get('project_id')
            ).first()

            if existing_entry:
                # Update existing entry
                for field, value in entry_data.items():
                    if value is not None:
                        setattr(existing_entry, field, value)
                updated_count += 1
            else:
                # Create new entry
                new_entry = _5G_Rollout_Sheet(**entry_data)
                db.add(new_entry)
                inserted_count += 1

        db.commit()

        # Create audit log
        await create_audit_log(
            db=db,
            user_id=current_user.id,
            action="bulk_upload_rollout_sheet",
            resource_type="rollout_sheet",
            resource_id="bulk",
            resource_name=file.filename,
            details=json.dumps({
                "inserted_count": inserted_count,
                "updated_count": updated_count,
                "filename": file.filename,
                "project_id": project_id
            }),
            ip_address=get_client_ip(request),
            user_agent=request.headers.get("User-Agent")
        )

        return UploadResponse(
            inserted=inserted_count,
            updated=updated_count,
            message=f"Successfully processed CSV: {inserted_count} inserted, {updated_count} updated"
        )
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error processing CSV: {str(e)}"
        )


@rolloutSheetRoute.delete("/rollout-sheet/delete-all/{project_id}")
async def delete_all_rollout_sheets_for_project(
        project_id: str,
        request: Request,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    """
    Delete all 5G Rollout Sheet entries for a specific project.
    Requires 'all' permission (senior_admin only).
    """
    # Check permission - only senior_admin can delete
    if not check_admin_access(current_user, "all"):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You are not authorized to delete rollout sheet entries. Contact the Senior Admin."
        )

    try:
        # Count entries to be deleted
        entries_count = db.query(_5G_Rollout_Sheet).filter(
            _5G_Rollout_Sheet.project_id == project_id
        ).count()

        if entries_count == 0:
            raise HTTPException(
                status_code=404,
                detail="No rollout sheet entries found for this project"
            )

        # Delete all entries for this project
        deleted_count = db.query(_5G_Rollout_Sheet).filter(
            _5G_Rollout_Sheet.project_id == project_id
        ).delete(synchronize_session=False)

        db.commit()

        # Create audit log
        await create_audit_log(
            db=db,
            user_id=current_user.id,
            action="delete_all_rollout_sheets",
            resource_type="rollout_sheet",
            resource_id=project_id,
            resource_name=f"Project {project_id}",
            details=json.dumps({
                "project_id": project_id,
                "deleted_count": deleted_count
            }),
            ip_address=get_client_ip(request),
            user_agent=request.headers.get("User-Agent")
        )

        return {
            "message": f"Successfully deleted all rollout sheet entries for project {project_id}",
            "deleted_count": deleted_count
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error deleting rollout sheet entries: {str(e)}"
        )


# ===========================
# BOQ GENERATION
# ===========================

def get_scope_column(scope: str) -> Optional[str]:
    """
    Map a scope value from the rollout sheet to the corresponding column in OD_BOQ_Item.
    Returns the first matching column field name, or None if no match.
    Handles scopes with newlines and extra whitespace.
    """
    if not scope:
        return None

    # Normalize whitespace: replace all whitespace (including newlines) with single space
    import re
    scope_normalized = re.sub(r'\s+', ' ', scope.strip()).lower()

    # Iterate through LEVEL1_CATEGORIES to find the first matching column
    for field_name, category_name in LEVEL1_CATEGORIES.items():
        category_normalized = re.sub(r'\s+', ' ', category_name.strip()).lower()
        if category_normalized == scope_normalized:
            return field_name

    # Try partial matching if exact match fails
    for field_name, category_name in LEVEL1_CATEGORIES.items():
        category_normalized = re.sub(r'\s+', ' ', category_name.strip()).lower()
        if scope_normalized in category_normalized or category_normalized in scope_normalized:
            return field_name

    return None


@rolloutSheetRoute.get("/rollout-sheet/{entry_id}/generate-boq")
def generate_boq_for_rollout_entry(
        entry_id: int,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    """
    Generate BOQ for a specific rollout sheet entry.

    Steps:
    1. Get the rollout sheet entry by ID to retrieve the scope
    2. Map scope to the corresponding column in OD_BOQ_Item
    3. Filter OD_BOQ_Item by project_id and get rows where that column value > 0
    4. Match descriptions with CustomerPO (exact match) to get line, item_job, quantity, price
    5. Return CSV-formatted string
    """
    # Get the rollout sheet entry
    entry = db.query(_5G_Rollout_Sheet).filter(_5G_Rollout_Sheet.id == entry_id).first()

    if not entry:
        raise HTTPException(status_code=404, detail="Rollout sheet entry not found")

    # Check if user has access to this entry's project
    if current_user.role.name != "senior_admin" and entry.project_id:
        user_access = db.query(UserProjectAccess).filter(
            UserProjectAccess.user_id == current_user.id,
            UserProjectAccess.project_id == entry.project_id
        ).first()

        if not user_access:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You do not have access to this rollout sheet entry."
            )

    # Get the scope and map it to the column
    scope = entry.scope
    if not scope:
        raise HTTPException(
            status_code=400,
            detail="This rollout entry has no scope defined. Cannot generate BOQ."
        )

    scope_column = get_scope_column(scope)
    if not scope_column:
        raise HTTPException(
            status_code=400,
            detail=f"Could not map scope '{scope}' to any BOQ column. Available scopes: {', '.join(set(LEVEL1_CATEGORIES.values()))}"
        )

    # Query OD_BOQ_Item for items with non-zero values in the scope column
    boq_items = db.query(ODBOQItem).filter(
        ODBOQItem.project_id == entry.project_id
    ).all()

    # Filter items where the scope column value is not null/zero
    filtered_items = []
    for item in boq_items:
        qty_value = getattr(item, scope_column, None)
        if qty_value is not None and qty_value != 0:
            # Override quantity based on category (service type)
            category_lower = (item.category or '').lower()
            description_lower = (item.description or '').lower()

            if category_lower == 'hw':
                # Hardware -> quantity = 1
                final_qty = 1
            elif category_lower == 'sw':
                # Software -> quantity = 3
                final_qty = 3
            elif category_lower == 'service':
                # Service -> quantity = 1, except if contains 'DU breaker' then 2
                if 'dc breaker' in description_lower:
                    final_qty = 2
                else:
                    final_qty = 1
            else:
                # If category doesn't match, use original quantity
                final_qty = qty_value

            filtered_items.append({
                'description': item.description,
                'uom': item.uom,
                'category': item.category,
                'bu': item.bu,
                'boq_qty': final_qty
            })

    if not filtered_items:
        raise HTTPException(
            status_code=404,
            detail=f"No BOQ items found with non-zero quantities for scope '{scope}' in this project."
        )

    # Get all CustomerPO items for this project for matching
    customer_po_items = db.query(CustomerPO).filter(
        CustomerPO.project_id == entry.project_id
    ).all()

    # Create a lookup dictionary by description (exact match)
    po_lookup = {}
    for po_item in customer_po_items:
        if po_item.description:
            po_lookup[po_item.description.strip()] = {
                'line': po_item.line,
                'item_job': po_item.item_job,
                'po_qty': po_item.quantity,
                'price': po_item.price
            }

    # Helper function to replace null/empty with space
    def safe_value(val):
        if val is None or val == '':
            return ' '
        return val

    # Build the result data by matching descriptions
    result_data = []
    for item in filtered_items:
        desc = item['description'].strip() if item['description'] else ' '
        po_info = po_lookup.get(desc.strip(), {})

        result_data.append({
            'Line': safe_value(po_info.get('line')),
            'Item/Job': safe_value(po_info.get('item_job')),
            'Description': safe_value(desc),
            'Category': safe_value(item['category']),
            'BU': safe_value(item['bu']),
            'UOM': safe_value(item['uom']),
            'BOQ Qty': safe_value(item['boq_qty']),
            'PO Qty': safe_value(po_info.get('po_qty')),
            'Price': safe_value(po_info.get('price'))
        })

    # Get project details for header
    project = db.query(DUProject).filter(DUProject.pid_po == entry.project_id).first()
    project_name = project.project_name if project else entry.project_id
    project_po = project.po if project else entry.project_id

    # Get current date
    from datetime import datetime
    current_date = datetime.now().strftime('%d-%b-%y')

    # Generate CSV string with metadata headers
    csv_lines = []

    # Add metadata header rows (3 rows)
    csv_lines.append(f'" "," "," "," ",DU BOQ," "," "," "," ",')
    csv_lines.append(f'BPO Number:,{project_po}," "," "," ",Date:,{current_date}," "," "')
    csv_lines.append(f'Scope Description:,{scope}," "," "," ",Site Classification,{entry.sps_category or "N/A"}," "," "')
    csv_lines.append(f'Vendor:,Nokia," "," "," ",Site ID:,{entry.site_id}," "," "')
    csv_lines.append(f'" "," "," "," "," "," "," "," "," ",') # Empty row separator

    # Add data table headers
    csv_headers = ['Line', 'Item/Job', 'Description', 'Category', 'BU', 'UOM', 'BOQ Qty', 'PO Qty', 'Price']
    csv_lines.append(','.join(csv_headers))

    # Add data rows
    for row in result_data:
        csv_row = []
        for header in csv_headers:
            value = row.get(header, ' ')
            # Convert to string and replace null/empty with space
            value_str = str(value) if value is not None else ' '
            if value_str == '' or value_str == 'None':
                value_str = ' '
            # Escape commas and quotes in values
            if ',' in value_str or '"' in value_str:
                value_str = f'"{value_str.replace(chr(34), chr(34)+chr(34))}"'
            csv_row.append(value_str)
        csv_lines.append(','.join(csv_row))

    return '\n'.join(csv_lines)


@rolloutSheetRoute.post("/rollout-sheet/bulk-generate-boq")
def bulk_generate_boq(
        request: dict,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    """
    Generate BOQs for multiple rollout sheet entries.

    Request body:
    {
        "entry_ids": [1, 2, 3, ...]
    }

    Returns:
    {
        "results": [
            {
                "entry_id": 1,
                "site_id": "SITE001",
                "scope": "New SRAN",
                "csv_content": "...",
                "success": true
            },
            {
                "entry_id": 2,
                "site_id": "SITE002",
                "error": "No BOQ items found",
                "success": false
            }
        ],
        "total_requested": 2,
        "successful": 1,
        "failed": 1
    }
    """
    # Extract entry IDs from request
    entry_ids = request.get('entry_ids', [])

    if not entry_ids:
        raise HTTPException(
            status_code=400,
            detail="No entry IDs provided. Please provide an array of entry_ids."
        )

    if not isinstance(entry_ids, list):
        raise HTTPException(
            status_code=400,
            detail="entry_ids must be an array of integers."
        )

    results = []
    successful = 0
    failed = 0

    # Process each entry
    for entry_id in entry_ids:
        site_id = None  # Initialize early for error reporting
        try:
            # Get the rollout sheet entry
            entry = db.query(_5G_Rollout_Sheet).filter(_5G_Rollout_Sheet.id == entry_id).first()

            if not entry:
                results.append({
                    "entry_id": entry_id,
                    "site_id": f"Unknown (Entry ID: {entry_id})",
                    "scope": None,
                    "error": "Rollout sheet entry not found",
                    "success": False
                })
                failed += 1
                continue

            # Get site_id early for error reporting
            site_id = entry.site_id

            # Check if user has access to this entry's project
            if current_user.role.name != "senior_admin" and entry.project_id:
                user_access = db.query(UserProjectAccess).filter(
                    UserProjectAccess.user_id == current_user.id,
                    UserProjectAccess.project_id == entry.project_id
                ).first()

                if not user_access:
                    results.append({
                        "entry_id": entry_id,
                        "site_id": entry.site_id,
                        "scope": entry.scope,
                        "error": "Access denied to this project",
                        "success": False
                    })
                    failed += 1
                    continue

            # Get the scope and map it to the column
            scope = entry.scope
            if not scope:
                results.append({
                    "entry_id": entry_id,
                    "site_id": entry.site_id,
                    "scope": None,
                    "error": "No scope defined for this entry",
                    "success": False
                })
                failed += 1
                continue

            # Map scope to column (use local helper function)
            scope_column = get_scope_column(scope)

            if not scope_column:
                results.append({
                    "entry_id": entry_id,
                    "site_id": entry.site_id,
                    "scope": scope,
                    "error": f"Could not map scope '{scope}' to any BOQ column",
                    "success": False
                })
                failed += 1
                continue

            # Query OD_BOQ_Item for items with non-zero values in the scope column
            boq_items = db.query(ODBOQItem).filter(
                ODBOQItem.project_id == entry.project_id
            ).all()

            # Filter items where the scope column value is not null/zero
            filtered_items = []
            for item in boq_items:
                qty_value = getattr(item, scope_column, None)
                if qty_value is not None and qty_value != 0:
                    # Override quantity based on category (service type)
                    category_lower = (item.category or '').lower()
                    description_lower = (item.description or '').lower()

                    if category_lower == 'hw':
                        # Hardware -> quantity = 1
                        final_qty = 1
                    elif category_lower == 'sw':
                        # Software -> quantity = 3
                        final_qty = 3
                    elif category_lower == 'service':
                        # Service -> quantity = 1, except if contains 'dc breaker' then 2
                        if 'dc breaker' in description_lower:
                            final_qty = 2
                        else:
                            final_qty = 1
                    else:
                        # If category doesn't match, use original quantity
                        final_qty = qty_value

                    filtered_items.append({
                        'description': item.description,
                        'uom': item.uom,
                        'category': item.category,
                        'bu': item.bu,
                        'boq_qty': final_qty
                    })

            if not filtered_items:
                results.append({
                    "entry_id": entry_id,
                    "site_id": entry.site_id,
                    "scope": scope,
                    "error": f"No BOQ items found with non-zero quantities for scope '{scope}'",
                    "success": False
                })
                failed += 1
                continue

            # Get all CustomerPO items for this project for matching
            customer_po_items = db.query(CustomerPO).filter(
                CustomerPO.project_id == entry.project_id
            ).all()

            # Create a lookup dictionary by description (exact match)
            po_lookup = {}
            for po_item in customer_po_items:
                if po_item.description:
                    po_lookup[po_item.description.strip()] = {
                        'line': po_item.line,
                        'item_job': po_item.item_job,
                        'po_qty': po_item.quantity,
                        'price': po_item.price
                    }

            # Helper function to replace null/empty with space
            def safe_value(val):
                if val is None or val == '':
                    return ' '
                return val

            # Build the result data by matching descriptions
            result_data = []
            for item in filtered_items:
                desc = item['description'].strip() if item['description'] else ' '
                po_info = po_lookup.get(desc.strip(), {})

                result_data.append({
                    'Line': safe_value(po_info.get('line')),
                    'Item/Job': safe_value(po_info.get('item_job')),
                    'Description': safe_value(desc),
                    'Category': safe_value(item['category']),
                    'BU': safe_value(item['bu']),
                    'UOM': safe_value(item['uom']),
                    'BOQ Qty': safe_value(item['boq_qty']),
                    'PO Qty': safe_value(po_info.get('po_qty')),
                    'Price': safe_value(po_info.get('price'))
                })

            # Get project details for header
            project = db.query(DUProject).filter(DUProject.pid_po == entry.project_id).first()
            project_name = project.project_name if project else entry.project_id
            project_po = project.po if project else entry.project_id

            # Get current date
            from datetime import datetime
            current_date = datetime.now().strftime('%d-%b-%y')

            # Generate CSV string with metadata headers
            csv_lines = []

            # Add metadata header rows (3 rows)
            csv_lines.append(f'" "," "," "," ",DU BOQ," "," "," "," ",')
            csv_lines.append(f'BPO Number:,{project_po}," "," "," ",Date:,{current_date}," "," "')
            csv_lines.append(f'Scope Description:,{scope}," "," "," ",Site Classification,{entry.sps_category or "N/A"}," "," "')
            csv_lines.append(f'Vendor:,Nokia," "," "," ",Site ID:,{entry.site_id}," "," "')
            csv_lines.append(f'" "," "," "," "," "," "," "," "," ",') # Empty row separator

            # Add data table headers
            csv_headers = ['Line', 'Item/Job', 'Description', 'Category', 'BU', 'UOM', 'BOQ Qty', 'PO Qty', 'Price']
            csv_lines.append(','.join(csv_headers))

            # Add data rows
            for row in result_data:
                csv_row = []
                for header in csv_headers:
                    value = row.get(header, ' ')
                    # Convert to string and replace null/empty with space
                    value_str = str(value) if value is not None else ' '
                    if value_str == '' or value_str == 'None':
                        value_str = ' '
                    # Escape commas and quotes in values
                    if ',' in value_str or '"' in value_str:
                        value_str = f'"{value_str.replace(chr(34), chr(34)+chr(34))}"'
                    csv_row.append(value_str)
                csv_lines.append(','.join(csv_row))

            csv_content = '\n'.join(csv_lines)

            results.append({
                "entry_id": entry_id,
                "site_id": entry.site_id,
                "scope": scope,
                "csv_content": csv_content,
                "success": True
            })
            successful += 1

        except Exception as e:
            # Catch any unexpected errors
            results.append({
                "entry_id": entry_id,
                "site_id": site_id or f"Unknown (Entry ID: {entry_id})",
                "scope": None,
                "error": str(e),
                "success": False
            })
            failed += 1

    return {
        "results": results,
        "total_requested": len(entry_ids),
        "successful": successful,
        "failed": failed
    }


# ===========================
# EXCEL DOWNLOAD USING TEMPLATE
# ===========================

def generate_boq_data_for_entry(entry_id: int, db: Session):
    """
    Generate BOQ data for a single rollout entry.
    Returns a dictionary with entry metadata and BOQ items.
    """
    # Get the rollout sheet entry
    entry = db.query(_5G_Rollout_Sheet).filter(_5G_Rollout_Sheet.id == entry_id).first()

    if not entry:
        return None

    # Get the scope and map it to the column
    scope = entry.scope
    if not scope:
        return None

    scope_column = get_scope_column(scope)
    if not scope_column:
        return None

    # Query OD_BOQ_Item for items with non-zero values in the scope column
    boq_items = db.query(ODBOQItem).filter(
        ODBOQItem.project_id == entry.project_id
    ).all()

    # Filter items where the scope column value is not null/zero
    filtered_items = []
    for item in boq_items:
        qty_value = getattr(item, scope_column, None)
        if qty_value is not None and qty_value != 0:
            # Override quantity based on category (service type)
            category_lower = (item.category or '').lower()
            description_lower = (item.description or '').lower()

            if category_lower == 'hw':
                final_qty = 1
            elif category_lower == 'sw':
                final_qty = 3
            elif category_lower == 'service':
                if 'dc breaker' in description_lower:
                    final_qty = 2
                else:
                    final_qty = 1
            else:
                final_qty = qty_value

            filtered_items.append({
                'description': item.description,
                'uom': item.uom,
                'category': item.category,
                'bu': item.bu,
                'boq_qty': final_qty
            })

    if not filtered_items:
        return None

    # Get all CustomerPO items for this project for matching
    customer_po_items = db.query(CustomerPO).filter(
        CustomerPO.project_id == entry.project_id
    ).all()

    # Create a lookup dictionary by description (exact match)
    po_lookup = {}
    for po_item in customer_po_items:
        if po_item.description:
            po_lookup[po_item.description.strip()] = {
                'line': po_item.line,
                'item_job': po_item.item_job,
                'po_qty': po_item.quantity,
                'price': po_item.price
            }

    # Build the result data by matching descriptions
    result_data = []
    for item in filtered_items:
        desc = item['description'].strip() if item['description'] else ''
        po_info = po_lookup.get(desc, {})

        result_data.append({
            'line': po_info.get('line'),
            'bu': item['bu'],
            'item_job': po_info.get('item_job'),  # ERP Item Code
            'description': desc,
            'budget_line': item['category'],  # Budget line from category
            'qty': item['boq_qty'],
            'unit_price': po_info.get('price'),
            'total_usd': '',  # Leave blank as per user request
            'total_aed': '',  # Leave blank as per user request
            'site_id_list': entry.site_id,
            'smp': entry.smp_number  # SMP from rollout sheet
        })

    # Get project details for header
    project = db.query(DUProject).filter(DUProject.pid_po == entry.project_id).first()
    project_po = project.po if project else entry.project_id

    return {
        'entry_id': entry.id,
        'site_id': entry.site_id,
        'scope': scope,
        'project_po': project_po,
        'sps_category': entry.sps_category or 'N/A',
        'data': result_data
    }


def create_excel_from_template(boq_entries: List[dict], template_path: str, is_bulk: bool = False):
    """
    Create an Excel file from the template with BOQ data.

    Args:
        boq_entries: List of BOQ entry dictionaries from generate_boq_data_for_entry
        template_path: Path to the Excel template file
        is_bulk: If True, combine all entries in one sheet; if False, use first entry only

    Returns:
        BytesIO object containing the Excel file
    """
    # Load the template
    wb = load_workbook(template_path)
    ws = wb.active

    # Save the footer rows (blue section with names) before clearing
    # The footer typically starts after the data section
    # We need to find where the data ends and footer begins
    max_row = ws.max_row
    footer_rows = []

    # Get a sample border style from the template (from header row)
    # Try row 9 (header) first, then row 10 (data) if needed
    template_border = None

    # Try to get border from header row (row 9)
    if ws.max_row >= 9:
        for col in range(2, 13):  # Check columns B through L
            sample_cell = ws.cell(row=9, column=col)
            if sample_cell.border and sample_cell.border.left and sample_cell.border.left.style:
                template_border = sample_cell.border.copy()
                break

    # If not found in header, try row 10 (data row)
    if template_border is None and max_row >= 10:
        for col in range(2, 13):
            sample_cell = ws.cell(row=10, column=col)
            if sample_cell.border and sample_cell.border.left and sample_cell.border.left.style:
                template_border = sample_cell.border.copy()
                break

    # If no border found in template, create a default thin black border
    if template_border is None:
        thin_side = Side(style='thin', color='000000')
        template_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # Look for the footer by checking for cells with background fill (blue color)
    # Usually the footer is at the bottom of the sheet
    if max_row >= 10:
        # Save rows that have formatting (likely footer) starting from a reasonable position
        # Check last 20 rows for formatted content
        footer_start = None
        for row_idx in range(max(10, max_row - 20), max_row + 1):
            # Check if this row has any cells with fill color (indicating footer)
            has_fill = False
            for col_idx in range(1, 15):  # Check first 14 columns
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb and cell.fill.start_color.rgb != '00000000':
                    has_fill = True
                    footer_start = row_idx
                    break
            if has_fill:
                break

        # Save footer rows if found
        if footer_start:
            for row_idx in range(footer_start, max_row + 1):
                row_data = []
                for col_idx in range(1, 15):  # Save first 14 columns
                    cell = ws.cell(row=row_idx, column=col_idx)
                    row_data.append({
                        'value': cell.value,
                        'font': cell.font.copy() if cell.font else None,
                        'fill': cell.fill.copy() if cell.fill else None,
                        'border': cell.border.copy() if cell.border else None,
                        'alignment': cell.alignment.copy() if cell.alignment else None
                    })
                footer_rows.append(row_data)

        # Delete all data rows from 10 to max_row
        ws.delete_rows(10, max_row - 9)

    if not is_bulk:
        # Single site - use the first entry
        if not boq_entries or len(boq_entries) == 0:
            return None

        entry = boq_entries[0]

        # Validate that entry has data
        if not entry.get('data') or len(entry['data']) == 0:
            return None

        # Update header information (based on template structure from row analysis)
        # Row 5 (index 4): BPO Number
        ws.cell(row=5, column=5).value = entry['project_po']

        # Row 5 (index 4): Date
        ws.cell(row=5, column=9).value = datetime.now().strftime('%d-%b-%y')

        # Row 6 (index 5): Scope Description
        ws.cell(row=6, column=5).value = entry['scope']

        # Row 6 (index 5): Site Classification - goes in merged cell I6:K6 (column 9)
        ws.cell(row=6, column=9).value = entry['sps_category']

        # Row 7 (index 6): Site ID - goes in merged cell I7:K7 (column 9)
        ws.cell(row=7, column=9).value = entry['site_id']

        # Data starts at row 10 (index 9)
        start_row = 10

        # Sort data by line number (None values go to end)
        # Using tuple sorting: (is_none, value) ensures None values are at the end
        sorted_data = sorted(entry['data'], key=lambda x: (x.get('line') is None, x.get('line') if x.get('line') is not None else 0))

        for idx, item in enumerate(sorted_data):
            row_num = start_row + idx

            # Column mapping based on template analysis:
            # Col 1 (A): Empty but needs border
            cell = ws.cell(row=row_num, column=1)
            cell.value = ''
            if template_border:
                cell.border = template_border

            # Col 2 (B): BPO Line No
            cell = ws.cell(row=row_num, column=2)
            cell.value = item.get('line')
            if template_border:
                cell.border = template_border

            # Col 3 (C): BU
            cell = ws.cell(row=row_num, column=3)
            cell.value = item.get('bu')
            if template_border:
                cell.border = template_border

            # Col 4 (D): ERP Item Code
            cell = ws.cell(row=row_num, column=4)
            cell.value = item.get('item_job')
            if template_border:
                cell.border = template_border

            # Col 5 (E): PO Item Description
            cell = ws.cell(row=row_num, column=5)
            cell.value = item.get('description')
            if template_border:
                cell.border = template_border

            # Col 6 (F): Budget line - leave empty as per user request
            cell = ws.cell(row=row_num, column=6)
            cell.value = ''  # Leave empty
            if template_border:
                cell.border = template_border

            # Col 7 (G): QTY
            cell = ws.cell(row=row_num, column=7)
            cell.value = item.get('qty')
            if template_border:
                cell.border = template_border

            # Col 8 (H): Unit Price
            cell = ws.cell(row=row_num, column=8)
            cell.value = item.get('unit_price')
            if template_border:
                cell.border = template_border

            # Col 9 (I): TOTAL PER LINE USD - leave blank
            cell = ws.cell(row=row_num, column=9)
            cell.value = item.get('total_usd')
            if template_border:
                cell.border = template_border

            # Col 10 (J): TOTAL PER LINE AED - leave blank
            cell = ws.cell(row=row_num, column=10)
            cell.value = item.get('total_aed')
            if template_border:
                cell.border = template_border

            # Col 11 (K): Site ID list
            cell = ws.cell(row=row_num, column=11)
            cell.value = item.get('site_id_list')
            if template_border:
                cell.border = template_border

            # Col 12 (L): SMP
            cell = ws.cell(row=row_num, column=12)
            cell.value = item.get('smp')
            if template_border:
                cell.border = template_border

        # Ensure the last row has proper bottom borders
        # This is necessary because Excel sometimes doesn't render the bottom border properly
        if len(sorted_data) > 0 and template_border:
            last_row_num = start_row + len(sorted_data) - 1
            # Create a new border with explicit bottom border
            bottom_border = Border(
                left=template_border.left if template_border.left else Side(style='thin', color='000000'),
                right=template_border.right if template_border.right else Side(style='thin', color='000000'),
                top=template_border.top if template_border.top else Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')  # Explicitly set bottom border
            )
            for col in range(1, 13):  # Columns A through L
                cell = ws.cell(row=last_row_num, column=col)
                # Apply the border with explicit bottom
                cell.border = bottom_border

        # Restore footer rows after data
        if footer_rows:
            footer_start_row = start_row + len(sorted_data) + 1  # Start after data with one blank row
            for footer_idx, footer_row_data in enumerate(footer_rows):
                row_num = footer_start_row + footer_idx
                for col_idx, cell_data in enumerate(footer_row_data):
                    cell = ws.cell(row=row_num, column=col_idx + 1)
                    cell.value = cell_data['value']
                    if cell_data['font']:
                        cell.font = cell_data['font']
                    if cell_data['fill']:
                        cell.fill = cell_data['fill']
                    if cell_data['border']:
                        cell.border = cell_data['border']
                    if cell_data['alignment']:
                        cell.alignment = cell_data['alignment']

    else:
        # Bulk mode - combine all entries in one sheet
        # Update header with first entry's project info
        if boq_entries and len(boq_entries) > 0:
            first_entry = boq_entries[0]
            ws.cell(row=5, column=5).value = first_entry['project_po']
            ws.cell(row=5, column=9).value = datetime.now().strftime('%d-%b-%y')
            ws.cell(row=6, column=5).value = "Multiple Scopes (Bulk)"
            ws.cell(row=6, column=9).value = "Multiple Sites"
            ws.cell(row=7, column=9).value = f"{len(boq_entries)} sites"

        # Collect all data from all entries and sort by line number
        all_data = []
        for entry in boq_entries:
            for item in entry['data']:
                # Add site_id to each item for reference
                item_with_site = item.copy()
                item_with_site['site_id_list'] = entry['site_id']
                item_with_site['smp'] = entry.get('smp')
                all_data.append(item_with_site)

        # Sort all data by line number (None values go to end)
        # Using tuple sorting: (is_none, value) ensures None values are at the end
        sorted_data = sorted(all_data, key=lambda x: (x.get('line') is None, x.get('line') if x.get('line') is not None else 0))

        # Data starts at row 10
        start_row = 10

        for idx, item in enumerate(sorted_data):
            row_num = start_row + idx

            # Same column mapping as single mode
            # Col 1 (A): Empty but needs border
            cell = ws.cell(row=row_num, column=1)
            cell.value = ''
            if template_border:
                cell.border = template_border

            cell = ws.cell(row=row_num, column=2)
            cell.value = item.get('line')
            if template_border:
                cell.border = template_border

            cell = ws.cell(row=row_num, column=3)
            cell.value = item.get('bu')
            if template_border:
                cell.border = template_border

            cell = ws.cell(row=row_num, column=4)
            cell.value = item.get('item_job')
            if template_border:
                cell.border = template_border

            cell = ws.cell(row=row_num, column=5)
            cell.value = item.get('description')
            if template_border:
                cell.border = template_border

            cell = ws.cell(row=row_num, column=6)
            cell.value = ''  # Leave budget line empty
            if template_border:
                cell.border = template_border

            cell = ws.cell(row=row_num, column=7)
            cell.value = item.get('qty')
            if template_border:
                cell.border = template_border

            cell = ws.cell(row=row_num, column=8)
            cell.value = item.get('unit_price')
            if template_border:
                cell.border = template_border

            cell = ws.cell(row=row_num, column=9)
            cell.value = item.get('total_usd')
            if template_border:
                cell.border = template_border

            cell = ws.cell(row=row_num, column=10)
            cell.value = item.get('total_aed')
            if template_border:
                cell.border = template_border

            cell = ws.cell(row=row_num, column=11)
            cell.value = item.get('site_id_list')
            if template_border:
                cell.border = template_border

            cell = ws.cell(row=row_num, column=12)
            cell.value = item.get('smp')
            if template_border:
                cell.border = template_border

        # Ensure the last row has proper bottom borders
        # This is necessary because Excel sometimes doesn't render the bottom border properly
        if len(sorted_data) > 0 and template_border:
            last_row_num = start_row + len(sorted_data) - 1
            # Create a new border with explicit bottom border
            bottom_border = Border(
                left=template_border.left if template_border.left else Side(style='thin', color='000000'),
                right=template_border.right if template_border.right else Side(style='thin', color='000000'),
                top=template_border.top if template_border.top else Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')  # Explicitly set bottom border
            )
            for col in range(1, 13):  # Columns A through L
                cell = ws.cell(row=last_row_num, column=col)
                # Apply the border with explicit bottom
                cell.border = bottom_border

        # Restore footer rows after data
        if footer_rows:
            footer_start_row = start_row + len(sorted_data) + 1  # Start after data with one blank row
            for footer_idx, footer_row_data in enumerate(footer_rows):
                row_num = footer_start_row + footer_idx
                for col_idx, cell_data in enumerate(footer_row_data):
                    cell = ws.cell(row=row_num, column=col_idx + 1)
                    cell.value = cell_data['value']
                    if cell_data['font']:
                        cell.font = cell_data['font']
                    if cell_data['fill']:
                        cell.fill = cell_data['fill']
                    if cell_data['border']:
                        cell.border = cell_data['border']
                    if cell_data['alignment']:
                        cell.alignment = cell_data['alignment']

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output


@rolloutSheetRoute.get("/rollout-sheet/{entry_id}/download-boq-excel")
async def download_single_boq_excel(
        entry_id: int,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    """
    Download a single BOQ as Excel file using the template.
    """
    # Get the rollout sheet entry
    entry = db.query(_5G_Rollout_Sheet).filter(_5G_Rollout_Sheet.id == entry_id).first()

    if not entry:
        raise HTTPException(status_code=404, detail="Rollout sheet entry not found")

    # Check if user has access to this entry's project
    if current_user.role.name != "senior_admin" and entry.project_id:
        user_access = db.query(UserProjectAccess).filter(
            UserProjectAccess.user_id == current_user.id,
            UserProjectAccess.project_id == entry.project_id
        ).first()

        if not user_access:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You do not have access to this rollout sheet entry."
            )

    # Generate BOQ data
    boq_data = generate_boq_data_for_entry(entry_id, db)

    if not boq_data:
        raise HTTPException(
            status_code=400,
            detail="Could not generate BOQ for this entry. Check that scope is defined and BOQ items exist."
        )

    # Get template path
    template_path = os.path.join(os.path.dirname(__file__), '..', '..', 'BOQ Formate.xlsx')

    if not os.path.exists(template_path):
        raise HTTPException(
            status_code=500,
            detail="BOQ template file not found"
        )

    # Generate Excel file
    excel_file = create_excel_from_template([boq_data], template_path, is_bulk=False)

    if not excel_file:
        raise HTTPException(
            status_code=500,
            detail="Failed to generate Excel file"
        )

    # Return as downloadable file
    filename = f"BOQ_{entry.site_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@rolloutSheetRoute.post("/rollout-sheet/bulk-download-boq-excel")
async def bulk_download_boq_excel(
        request: dict,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    """
    Download multiple BOQs combined in one Excel file using the template.
    All sites will be in one sheet, ordered by line number.

    Request body:
    {
        "entry_ids": [1, 2, 3, ...]
    }
    """
    # Extract entry IDs from request
    entry_ids = request.get('entry_ids', [])

    if not entry_ids:
        raise HTTPException(
            status_code=400,
            detail="No entry IDs provided. Please provide an array of entry_ids."
        )

    if not isinstance(entry_ids, list):
        raise HTTPException(
            status_code=400,
            detail="entry_ids must be an array of integers."
        )

    # Generate BOQ data for all entries
    boq_entries = []
    failed_entries = []

    for entry_id in entry_ids:
        try:
            # Check access for each entry
            entry = db.query(_5G_Rollout_Sheet).filter(_5G_Rollout_Sheet.id == entry_id).first()

            if not entry:
                failed_entries.append({"entry_id": entry_id, "reason": "Entry not found"})
                continue

            # Check if user has access to this entry's project
            if current_user.role.name != "senior_admin" and entry.project_id:
                user_access = db.query(UserProjectAccess).filter(
                    UserProjectAccess.user_id == current_user.id,
                    UserProjectAccess.project_id == entry.project_id
                ).first()

                if not user_access:
                    failed_entries.append({"entry_id": entry_id, "reason": "Access denied"})
                    continue

            # Generate BOQ data
            boq_data = generate_boq_data_for_entry(entry_id, db)

            if boq_data:
                boq_entries.append(boq_data)
            else:
                failed_entries.append({"entry_id": entry_id, "reason": "No BOQ data available"})

        except Exception as e:
            failed_entries.append({"entry_id": entry_id, "reason": str(e)})

    if not boq_entries:
        raise HTTPException(
            status_code=400,
            detail=f"Could not generate BOQ for any of the selected entries. Failures: {failed_entries}"
        )

    # Get template path
    template_path = os.path.join(os.path.dirname(__file__), '..', '..', 'BOQ Formate.xlsx')

    if not os.path.exists(template_path):
        raise HTTPException(
            status_code=500,
            detail="BOQ template file not found"
        )

    # Generate Excel file with all entries
    excel_file = create_excel_from_template(boq_entries, template_path, is_bulk=True)

    if not excel_file:
        raise HTTPException(
            status_code=500,
            detail="Failed to generate Excel file"
        )

    # Return as downloadable file
    filename = f"BOQ_Bulk_{len(boq_entries)}_sites_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ===========================
# EXCEL DOWNLOAD FROM CSV DATA (WITH MODIFICATIONS)
# ===========================

def parse_csv_data_to_boq_format(csv_data: List[List[str]], site_id: str, entry_id: int, db: Session):
    """
    Parse CSV data from frontend and convert to BOQ format for Excel generation.

    Args:
        csv_data: List of lists representing CSV rows (includes metadata rows 0-4, headers at row 5, data from row 6+)
        site_id: Site ID for the BOQ
        entry_id: Entry ID to fetch metadata from database
        db: Database session

    Returns:
        Dictionary with entry metadata and BOQ data
    """
    if not csv_data or len(csv_data) < 2:  # Need at least headers + 1 data row
        return None

    # Get the rollout sheet entry for metadata
    entry = db.query(_5G_Rollout_Sheet).filter(_5G_Rollout_Sheet.id == entry_id).first()
    if not entry:
        return None

    # Get project details for header
    project = db.query(DUProject).filter(DUProject.pid_po == entry.project_id).first()
    project_po = project.po if project else entry.project_id

    # The CSV structure from generate_boq_for_rollout_entry is:
    # Row 0-4: Metadata headers
    # Row 5: Data table headers ['Line', 'Item/Job', 'Description', 'Category', 'BU', 'UOM', 'BOQ Qty', 'PO Qty', 'Price']
    # Row 6+: Data rows

    # Find the header row (look for 'Line' column)
    header_row_idx = None
    for idx in range(min(6, len(csv_data))):  # Check first 6 rows
        row = csv_data[idx]
        if any('line' in str(cell).lower() and 'site' not in str(cell).lower() for cell in row):
            header_row_idx = idx
            break

    if header_row_idx is None:
        # Fallback: assume row 5 is headers if not found
        header_row_idx = 5 if len(csv_data) > 5 else 0

    # Parse CSV headers to find column indices
    headers = csv_data[header_row_idx]

    # Common header names to look for (case-insensitive)
    header_mapping = {}
    for idx, header in enumerate(headers):
        header_lower = str(header).lower().strip()
        if 'line' in header_lower and 'site' not in header_lower:  # Match 'Line' or 'BPO Line' but not 'Site ID'
            if 'line' not in header_mapping:  # Only map the first line column
                header_mapping['line'] = idx
        elif 'bu' == header_lower or header_lower.startswith('bu ') or header_lower.endswith(' bu'):
            header_mapping['bu'] = idx
        elif 'item/job' in header_lower or 'erp' in header_lower or 'item code' in header_lower:
            header_mapping['item_job'] = idx
        elif 'description' in header_lower and 'scope' not in header_lower:
            header_mapping['description'] = idx
        elif 'category' in header_lower or 'budget' in header_lower:
            header_mapping['budget_line'] = idx
        elif 'boq' in header_lower and 'qty' in header_lower:
            header_mapping['qty'] = idx
        elif ('unit price' in header_lower or header_lower == 'price') and 'total' not in header_lower:
            header_mapping['unit_price'] = idx
        elif 'total' in header_lower and 'usd' in header_lower:
            header_mapping['total_usd'] = idx
        elif 'total' in header_lower and 'aed' in header_lower:
            header_mapping['total_aed'] = idx
        elif 'site' in header_lower and 'id' in header_lower:
            header_mapping['site_id_list'] = idx
        elif 'smp' in header_lower:
            header_mapping['smp'] = idx

    # Parse data rows (start from row after headers)
    result_data = []
    data_start_row = header_row_idx + 1
    for row_idx in range(data_start_row, len(csv_data)):
        row = csv_data[row_idx]

        # Skip empty rows
        if not any(str(cell).strip() for cell in row):
            continue

        # Extract values based on header mapping
        def get_value(key, default=''):
            if key in header_mapping and header_mapping[key] < len(row):
                value = row[header_mapping[key]]
                return value if value is not None and str(value).strip() != '' else default
            return default

        # Parse line number as integer
        line_str = get_value('line', '')
        try:
            line_num = int(float(line_str)) if line_str else None
        except (ValueError, TypeError):
            line_num = None

        # Parse quantity as float
        qty_str = get_value('qty', '')
        try:
            qty = float(qty_str) if qty_str else None
        except (ValueError, TypeError):
            qty = None

        # Parse price as float
        price_str = get_value('unit_price', '')
        try:
            price = float(price_str) if price_str else None
        except (ValueError, TypeError):
            price = None

        result_data.append({
            'line': line_num,
            'bu': get_value('bu'),
            'item_job': get_value('item_job'),
            'description': get_value('description'),
            'budget_line': get_value('budget_line'),
            'qty': qty,
            'unit_price': price,
            'total_usd': get_value('total_usd'),
            'total_aed': get_value('total_aed'),
            'site_id_list': get_value('site_id_list', site_id),
            'smp': get_value('smp', entry.smp_number or '')
        })

    # Validate we have at least some data
    if not result_data:
        return None

    # Sort by line number (None values go to end)
    # Using tuple sorting: (is_none, value) ensures None values are at the end
    result_data.sort(key=lambda x: (x['line'] is None, x['line'] if x['line'] is not None else 0))

    return {
        'entry_id': entry.id,
        'site_id': site_id,
        'scope': entry.scope or '',
        'project_po': project_po,
        'sps_category': entry.sps_category or 'N/A',
        'data': result_data
    }


@rolloutSheetRoute.post("/rollout-sheet/download-boq-excel-from-csv")
async def download_boq_excel_from_csv(
        request: dict,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    """
    Download a single BOQ as Excel file using provided CSV data (with user modifications).

    Request body:
    {
        "entry_id": 123,
        "site_id": "SITE001",
        "csv_data": [
            ["Line", "BU", "Description", ...],  # headers
            ["1", "RAN", "Item 1", ...],         # data rows
            ...
        ]
    }
    """
    entry_id = request.get('entry_id')
    site_id = request.get('site_id')
    csv_data = request.get('csv_data')

    if not entry_id or not site_id or not csv_data:
        raise HTTPException(
            status_code=400,
            detail="Missing required fields: entry_id, site_id, and csv_data are required."
        )

    # Check if user has access to this entry
    entry = db.query(_5G_Rollout_Sheet).filter(_5G_Rollout_Sheet.id == entry_id).first()

    if not entry:
        raise HTTPException(status_code=404, detail="Rollout sheet entry not found")

    if current_user.role.name != "senior_admin" and entry.project_id:
        user_access = db.query(UserProjectAccess).filter(
            UserProjectAccess.user_id == current_user.id,
            UserProjectAccess.project_id == entry.project_id
        ).first()

        if not user_access:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You do not have access to this rollout sheet entry."
            )

    # Parse CSV data and convert to BOQ format
    boq_data = parse_csv_data_to_boq_format(csv_data, site_id, entry_id, db)

    if not boq_data:
        raise HTTPException(
            status_code=400,
            detail="Could not parse CSV data. Ensure data is properly formatted."
        )

    # Get template path
    template_path = os.path.join(os.path.dirname(__file__), '..', '..', 'BOQ Formate.xlsx')

    if not os.path.exists(template_path):
        raise HTTPException(
            status_code=500,
            detail="BOQ template file not found"
        )

    # Generate Excel file
    excel_file = create_excel_from_template([boq_data], template_path, is_bulk=False)

    if not excel_file:
        raise HTTPException(
            status_code=500,
            detail="Failed to generate Excel file"
        )

    # Return as downloadable file
    filename = f"BOQ_{site_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@rolloutSheetRoute.post("/rollout-sheet/bulk-download-boq-excel-from-csv")
async def bulk_download_boq_excel_from_csv(
        request: dict,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    """
    Download multiple BOQs combined in one Excel file using provided CSV data (with user modifications).

    Request body:
    {
        "boq_data_list": [
            {
                "entry_id": 123,
                "site_id": "SITE001",
                "csv_data": [[...], [...], ...]
            },
            {
                "entry_id": 124,
                "site_id": "SITE002",
                "csv_data": [[...], [...], ...]
            },
            ...
        ]
    }
    """
    boq_data_list = request.get('boq_data_list', [])

    if not boq_data_list:
        raise HTTPException(
            status_code=400,
            detail="No BOQ data provided. Please provide boq_data_list."
        )

    # Parse all BOQ data
    parsed_boq_entries = []
    failed_entries = []

    for boq_item in boq_data_list:
        try:
            entry_id = boq_item.get('entry_id')
            site_id = boq_item.get('site_id')
            csv_data = boq_item.get('csv_data')

            if not entry_id or not site_id or not csv_data:
                failed_entries.append({
                    "entry_id": entry_id,
                    "reason": "Missing entry_id, site_id, or csv_data"
                })
                continue

            # Check access
            entry = db.query(_5G_Rollout_Sheet).filter(_5G_Rollout_Sheet.id == entry_id).first()

            if not entry:
                failed_entries.append({"entry_id": entry_id, "reason": "Entry not found"})
                continue

            if current_user.role.name != "senior_admin" and entry.project_id:
                user_access = db.query(UserProjectAccess).filter(
                    UserProjectAccess.user_id == current_user.id,
                    UserProjectAccess.project_id == entry.project_id
                ).first()

                if not user_access:
                    failed_entries.append({"entry_id": entry_id, "reason": "Access denied"})
                    continue

            # Parse CSV data
            boq_data = parse_csv_data_to_boq_format(csv_data, site_id, entry_id, db)

            if boq_data:
                parsed_boq_entries.append(boq_data)
            else:
                failed_entries.append({
                    "entry_id": entry_id,
                    "reason": "Could not parse CSV data"
                })

        except Exception as e:
            failed_entries.append({"entry_id": boq_item.get('entry_id'), "reason": str(e)})

    if not parsed_boq_entries:
        raise HTTPException(
            status_code=400,
            detail=f"Could not parse any BOQ data. Failures: {failed_entries}"
        )

    # Get template path
    template_path = os.path.join(os.path.dirname(__file__), '..', '..', 'BOQ Formate.xlsx')

    if not os.path.exists(template_path):
        raise HTTPException(
            status_code=500,
            detail="BOQ template file not found"
        )

    # Generate Excel file with all entries
    excel_file = create_excel_from_template(parsed_boq_entries, template_path, is_bulk=True)

    if not excel_file:
        raise HTTPException(
            status_code=500,
            detail="Failed to generate Excel file"
        )

    # Return as downloadable file
    filename = f"BOQ_Bulk_{len(parsed_boq_entries)}_sites_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
